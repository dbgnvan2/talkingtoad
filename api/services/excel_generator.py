import io
import logging
from openpyxl import Workbook
from openpyxl.styles import Font, Alignment, PatternFill
from api.models.job import CrawlJob
from api.models.issue import Issue

logger = logging.getLogger(__name__)

def generate_excel_report(
    job: CrawlJob,
    issues: list[Issue],
    summary: dict,
    image_summary: dict = None,
    images: list = None,
    performance: dict | None = None,
    priority_pages: list[dict] | None = None,
    prevalence: list | None = None,
    priority_pages_for_roadmap: list[dict] | None = None,
) -> bytes:
    """Generate a multi-sheet Excel workbook from crawl data."""
    wb = Workbook()
    
    # ── Summary Sheet ──────────────────────────────────────────────────────
    ws_summary = wb.active
    ws_summary.title = "Summary"
    
    # Styling
    header_font = Font(bold=True, size=14)
    label_font = Font(bold=True)
    
    ws_summary["A1"] = "TalkingToad SEO Audit"
    ws_summary["A1"].font = header_font
    
    ws_summary["A3"] = "Target URL:"
    ws_summary["B3"] = job.target_url
    ws_summary["A3"].font = label_font
    
    ws_summary["A4"] = "Health Score:"
    ws_summary["B4"] = summary.get("health_score", 0)
    ws_summary["A4"].font = label_font

    ws_summary["A5"] = "Agent Health Score:"
    ws_summary["B5"] = summary.get("agent_health_score", 0)
    ws_summary["A5"].font = label_font

    ws_summary["A6"] = "Total Pages:"
    ws_summary["B6"] = summary.get("pages_crawled", 0)
    ws_summary["A6"].font = label_font

    ws_summary["A7"] = "Total Issues:"
    ws_summary["B7"] = summary.get("total_issues", 0)
    ws_summary["A7"].font = label_font

    # Category totals table
    ws_summary["A9"] = "Issues by Category"
    ws_summary["A9"].font = Font(bold=True, size=12)
    
    ws_summary.append([]) # spacer
    ws_summary.append(["Category", "Count"])
    # Bold the mini-header
    for cell in ws_summary[ws_summary.max_row]:
        cell.font = label_font
        cell.fill = PatternFill(start_color="F3F4F6", end_color="F3F4F6", fill_type="solid")

    for cat, count in summary.get("by_category", {}).items():
        if count > 0:
            ws_summary.append([cat.replace('_', ' ').title(), count])

    # Adjust widths
    ws_summary.column_dimensions['A'].width = 25
    ws_summary.column_dimensions['B'].width = 50

    # ── AI Readiness Sheet ────────────────────────────────────────────────
    ws_ai = wb.create_sheet(title="AI Readiness")
    ws_ai["A1"] = "AI Readiness Report"
    ws_ai["A1"].font = header_font
    
    ws_ai["A3"] = "Live /llms.txt status:"
    ws_ai["A3"].font = label_font
    
    # Check if we have LLMS_TXT_MISSING in issues
    is_missing = any(i.issue_code == "LLMS_TXT_MISSING" for i in issues)
    ws_ai["B3"] = "MISSING" if is_missing else "FOUND"
    
    ws_ai["A5"] = "Proposed /llms.txt Content:"
    ws_ai["A5"].font = label_font
    
    # Put content in a single cell, wrap it
    proposed = job.llms_txt_custom or "Not generated yet."
    ws_ai["A6"] = proposed
    ws_ai["A6"].alignment = Alignment(wrap_text=True, vertical="top")
    ws_ai.merge_cells("A6:E20") # Give it some space
    
    ws_ai.column_dimensions['A'].width = 30
    ws_ai.column_dimensions['B'].width = 50
    ws_ai.column_dimensions['C'].width = 20
    ws_ai.column_dimensions['D'].width = 60
    ws_ai.column_dimensions['E'].width = 60

    # AI Readiness issue table with Confidence column
    ai_readiness_issues = [i for i in issues if i.category == "ai_readiness"]
    if ai_readiness_issues:
        start_row = 22  # after merged llms.txt block (A6:E20) + gap row
        headers = ["Code", "Severity", "Confidence", "Page URL", "Description"]
        for col_idx, header in enumerate(headers, 1):
            cell = ws_ai.cell(row=start_row, column=col_idx, value=header)
            cell.font = Font(bold=True)
            cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")

        for row_offset, issue in enumerate(ai_readiness_issues, 1):
            row = start_row + row_offset
            ws_ai.cell(row=row, column=1, value=issue.issue_code or "")
            ws_ai.cell(row=row, column=2, value=(issue.severity or "").upper())
            ws_ai.cell(row=row, column=3, value=issue.confidence_label or "")
            ws_ai.cell(row=row, column=4, value=issue.page_url or "")
            ws_ai.cell(row=row, column=5, value=issue.description or "")

    # ── Roadmap Sheet (E7.1) ───────────────────────────────────────────────
    # Same columns as the PDF roadmap: every surface offers the option, or the
    # decision not to offer it there is a recorded one (P25).
    if issues:
        try:
            from api.services.remediation import build_roadmap

            # Uncapped: the PDF points the reader here for the full list, so it
            # has to hold it. A cap here would make that promise false (F2/F6).
            roadmap_items, _weighted, _totals = build_roadmap(
                issues, prevalence=prevalence,
                priority_pages=priority_pages_for_roadmap or priority_pages,
                limit_per_phase=10**6,
            )
        except Exception:
            roadmap_items = []
        if roadmap_items:
            ws_road = wb.create_sheet(title="Roadmap")
            ws_road["A1"] = "Remediation Roadmap"
            ws_road["A1"].font = header_font
            headers = ["Phase", "Issue", "Code", "Category", "Owner", "Impact",
                       "Effort", "Pages affected", "Done when"]
            for col, h in enumerate(headers, 1):
                ws_road.cell(row=3, column=col, value=h).font = label_font
            row = 4
            for item in roadmap_items:
                values = [item.phase_title, item.title, item.code, item.category,
                          item.owner, item.impact, item.effort_label,
                          item.pages_affected, item.done_when]
                for col, v in enumerate(values, 1):
                    ws_road.cell(row=row, column=col, value=v)
                row += 1

    # ── Web Vitals Sheet (D2) ──────────────────────────────────────────────
    _vitals = getattr(job, "web_vitals", None)
    if _vitals and _vitals.get("rows"):
        ws_cwv = wb.create_sheet(title="Web Vitals")
        ws_cwv["A1"] = "Core Web Vitals"
        ws_cwv["A1"].font = header_font
        ws_cwv["A2"] = ("FIELD = 75th percentile across real Chrome users, last 28 days. "
                        "LAB = one synthetic run. Only field data raises a finding.")
        headers = ["URL", "Source", "LCP (ms)", "INP (ms)", "CLS",
                   "Lighthouse score", "Not measured because"]
        for col, h in enumerate(headers, 1):
            ws_cwv.cell(row=4, column=col, value=h).font = label_font
        row = 5
        for r in _vitals["rows"]:
            values = [r.get("url"), r.get("source"), r.get("lcp_ms"), r.get("inp_ms"),
                      r.get("cls"), r.get("performance_score"),
                      r.get("unavailable_reason")]
            for col, v in enumerate(values, 1):
                ws_cwv.cell(row=row, column=col, value=v)
            row += 1
        ws_cwv.column_dimensions['A'].width = 70
        ws_cwv.column_dimensions['G'].width = 50

    # ── Prevalence Sheet (E4.3) ────────────────────────────────────────────
    # How much of the indexable estate each defect touches. Scoring is unchanged;
    # this is the breadth lens. Spec: docs/pending/2026-08-29_E4-site-prevalence-escalation.md
    if prevalence:
        from api.services.prevalence import site_hygiene_score

        ws_prev = wb.create_sheet(title="Prevalence")
        ws_prev["A1"] = "Site Prevalence"
        ws_prev["A1"].font = header_font
        ws_prev["A2"] = "Site Hygiene:"
        ws_prev["B2"] = site_hygiene_score(prevalence)
        ws_prev["A2"].font = label_font
        ws_prev["A3"] = "Systemic defects:"
        ws_prev["B3"] = sum(1 for p in prevalence if p.tier == "systemic")
        ws_prev["A3"].font = label_font

        headers = ["Issue", "Code", "Category", "Severity", "Pages affected",
                   "Indexable pages", "Share", "Tier"]
        for col, h in enumerate(headers, 1):
            ws_prev.cell(row=5, column=col, value=h).font = label_font
        row = 6
        for p in prevalence:
            values = [p.human_description, p.code, p.category, p.severity,
                      p.pages_affected, p.indexable_pages, round(p.share, 4), p.tier]
            for col, v in enumerate(values, 1):
                ws_prev.cell(row=row, column=col, value=v)
            row += 1

    # ── Performance Sheet (E3.4) ───────────────────────────────────────────
    # Excel parity with the PDF: the same ranking must be reachable from every
    # export surface, not just the one that happened to be wired first (P25).
    # Spec: docs/pending/2026-08-29_E3-performance-data-in-report.md#E3.4
    if performance:
        ws_perf = wb.create_sheet(title="Performance")
        ws_perf["A1"] = "Search Performance"
        ws_perf["A1"].font = header_font
        ws_perf["A2"] = performance.get("source", "")
        ws_perf["A3"] = "Periods:"
        ws_perf["B3"] = ", ".join(performance.get("periods") or [])
        ws_perf["A3"].font = label_font
        if performance.get("is_stale"):
            ws_perf["A4"] = "Data age (days):"
            ws_perf["B4"] = performance.get("data_age_days")
            ws_perf["A4"].font = label_font

        site_rows = [
            ("Impressions", performance.get("total_impressions", 0)),
            ("Clicks", performance.get("total_clicks", 0)),
            ("Average CTR", performance.get("site_ctr", 0.0)),
            ("GA4 sessions", performance.get("total_sessions", 0)),
            ("Conversions", performance.get("total_conversions", 0)),
            ("AI-assistant sessions", performance.get("total_ai_referral_sessions", 0)),
            ("Pages with data", performance.get("pages_with_data", 0)),
        ]
        row = 6
        for label, value in site_rows:
            ws_perf.cell(row=row, column=1, value=label).font = label_font
            ws_perf.cell(row=row, column=2, value=value)
            row += 1

        row += 1
        shown = len(performance.get("top_by_impressions") or [])
        with_data = performance.get("pages_with_data", shown)
        label = (f"Top {shown} pages by impressions (of {with_data} pages with data)"
                 if with_data > shown else "Pages by impressions")
        ws_perf.cell(row=row, column=1, value=label).font = label_font
        row += 1
        headers = ["URL", "Impressions", "Clicks", "CTR", "Avg position",
                   "Sessions", "Conversions", "Health", "Period"]
        for col, h in enumerate(headers, 1):
            ws_perf.cell(row=row, column=col, value=h).font = label_font
        row += 1
        # NOTE: this list arrives already capped at `top_n` by
        # build_performance_summary. The count below says so rather than letting
        # the sheet imply completeness (rule 6) — an earlier comment here claimed
        # "uncapped on purpose", which was simply untrue.
        for r in performance.get("top_by_impressions") or []:
            for col, key in enumerate(
                ["url", "impressions", "clicks", "ctr", "position",
                 "sessions", "conversions", "health_score", "period"], 1
            ):
                ws_perf.cell(row=row, column=col, value=r.get(key))
            row += 1

        if performance.get("low_ctr_high_impression"):
            row += 1
            ws_perf.cell(row=row, column=1,
                         value="Seen but not clicked (above-median impressions, "
                               "below-average CTR)").font = label_font
            row += 1
            for col, h in enumerate(headers, 1):
                ws_perf.cell(row=row, column=col, value=h).font = label_font
            row += 1
            for r in performance["low_ctr_high_impression"]:
                for col, key in enumerate(
                    ["url", "impressions", "clicks", "ctr", "position",
                     "sessions", "conversions", "health_score", "period"], 1
                ):
                    ws_perf.cell(row=row, column=col, value=r.get(key))
                row += 1

    # ── Priority Pages Sheet (E3.4) ────────────────────────────────────────
    if priority_pages:
        ws_pri = wb.create_sheet(title="Priority Pages")
        ws_pri["A1"] = "Page Priority Work Queue"
        ws_pri["A1"].font = header_font
        headers = ["Rank", "URL", "Bucket", "Health", "Citability",
                   "Clicks", "Impressions", "CTR", "Position", "Conversions",
                   "Review flags"]
        for col, h in enumerate(headers, 1):
            ws_pri.cell(row=3, column=col, value=h).font = label_font
        row = 4
        for r in priority_pages:
            gsc = r.get("gsc") or {}
            flag = r.get("review_flag") or {}
            reasons = flag.get("reasons") if isinstance(flag, dict) else getattr(flag, "reasons", None)
            values = [
                r.get("priority_rank"), r.get("url"), r.get("bucket"),
                r.get("health_score"), r.get("citability_grade"),
                gsc.get("clicks"), gsc.get("impressions"), gsc.get("ctr"),
                gsc.get("position"), gsc.get("conversions"),
                "; ".join(reasons) if reasons else "",
            ]
            for col, v in enumerate(values, 1):
                ws_pri.cell(row=row, column=col, value=v)
            row += 1

    # ── Images Sheet ───────────────────────────────────────────────────────
    if image_summary and image_summary.get("total_images", 0) > 0:
        ws_img = wb.create_sheet(title="Images")
        ws_img["A1"] = "Image Health Report"
        ws_img["A1"].font = header_font

        ws_img["A3"] = "Image Health Score:"
        ws_img["B3"] = f"{image_summary.get('image_health_score', 0)}%"
        ws_img["A3"].font = label_font

        # E1.4 (P25/rule 6): the coverage disclosure must reach every surface,
        # not just the PDF — a bare "Total Images: 150" over a 1,284-image site
        # is the exact shape E1 was fixing.
        seen = getattr(job, "images_seen_total", None)
        collected = getattr(job, "images_collected", None)
        if seen and collected is not None and seen > collected:
            ws_img["D3"] = f"Coverage: analysed {collected} of {seen} images found"
            ws_img["D3"].font = label_font

        ws_img["A4"] = "Total Images:"
        ws_img["B4"] = image_summary.get("total_images", 0)
        ws_img["A4"].font = label_font

        ws_img["A5"] = "Total Size:"
        ws_img["B5"] = f"{image_summary.get('total_size_kb', 0)} KB"
        ws_img["A5"].font = label_font

        ws_img["A6"] = "Avg Load Time:"
        ws_img["B6"] = f"{image_summary.get('avg_load_time_ms', 0)}ms"
        ws_img["A6"].font = label_font

        # Format breakdown
        ws_img["A8"] = "Images by Format"
        ws_img["A8"].font = Font(bold=True, size=12)
        row_num = 9
        for fmt, count in sorted(image_summary.get("by_format", {}).items(), key=lambda x: -x[1]):
            ws_img[f"A{row_num}"] = fmt.upper()
            ws_img[f"B{row_num}"] = count
            row_num += 1

        # Top issues
        row_num += 1
        ws_img[f"A{row_num}"] = "Top Image Issues"
        ws_img[f"A{row_num}"].font = Font(bold=True, size=12)
        row_num += 1
        for code, count in sorted(image_summary.get("by_issue", {}).items(), key=lambda x: -x[1])[:10]:
            ws_img[f"A{row_num}"] = code.replace('IMG_', '').replace('_', ' ').title()
            ws_img[f"B{row_num}"] = count
            row_num += 1

        ws_img.column_dimensions['A'].width = 30
        ws_img.column_dimensions['B'].width = 20

        # Image details sheet
        if images and len(images) > 0:
            ws_img_list = wb.create_sheet(title="Image Details")
            img_headers = ["Score", "Filename", "URL", "Alt Text", "Size (KB)", "Dimensions", "Format", "Load Time (ms)", "Issues"]
            ws_img_list.append(img_headers)

            for cell in ws_img_list[1]:
                cell.font = label_font
                cell.fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
                cell.alignment = Alignment(horizontal="center")

            for img in images:
                # Normalize to dict if ImageInfo object
                d = img.to_dict() if hasattr(img, 'to_dict') else img if isinstance(img, dict) else {}
                score = d.get('overall_score', 0)
                filename = d.get('filename', '')
                url = d.get('url', '')
                alt = d.get('alt', '')
                size_bytes = d.get('file_size_bytes')
                size_kb = round(size_bytes / 1024, 1) if size_bytes else 0
                width = d.get('width', 0)
                height = d.get('height', 0)
                fmt = d.get('format', '')
                load_time = d.get('load_time_ms', 0)
                issues_list = d.get('issues', [])

                dimensions = f"{width}x{height}" if width and height else ""
                issues_str = ", ".join(issues_list) if issues_list else ""

                ws_img_list.append([
                    round(score) if score else 0,
                    filename or "",
                    url or "",
                    alt or "",
                    size_kb or 0,
                    dimensions,
                    (fmt or "").upper(),
                    load_time or 0,
                    issues_str
                ])

            ws_img_list.column_dimensions['A'].width = 8
            ws_img_list.column_dimensions['B'].width = 30
            ws_img_list.column_dimensions['C'].width = 60
            ws_img_list.column_dimensions['D'].width = 40
            ws_img_list.column_dimensions['E'].width = 12
            ws_img_list.column_dimensions['F'].width = 15
            ws_img_list.column_dimensions['G'].width = 10
            ws_img_list.column_dimensions['H'].width = 15
            ws_img_list.column_dimensions['I'].width = 50

            ws_img_list.auto_filter.ref = ws_img_list.dimensions

    # ── Issue Sheets (by Category) ─────────────────────────────────────────
    # Group issues by category
    from collections import defaultdict
    by_cat = defaultdict(list)
    for i in issues:
        by_cat[i.category].append(i)

    header_fill = PatternFill(start_color="E5E7EB", end_color="E5E7EB", fill_type="solid")
    
    # Sort categories for consistent tab order
    sorted_cats = sorted(by_cat.keys())
    
    for cat in sorted_cats:
        # Excel titles must be < 31 chars
        sheet_name = cat.replace('_', ' ').title()[:30]
        ws = wb.create_sheet(title=sheet_name)
        
        # E2: linking pages travel with broken-link rows so the "full list is in
        # the spreadsheet export" line in the UI and the PDF is actually true.
        headers = ["Severity", "URL", "Issue Code", "Description", "Recommendation",
                   "What to look for", "Linking pages", "Linking pages (total)"]
        ws.append(headers)
        
        # Style headers
        for cell in ws[1]:
            cell.font = label_font
            cell.fill = header_fill
            cell.alignment = Alignment(horizontal="center")

        from api.services.issue_evidence import evidence_for_excel

        for issue in by_cat[cat]:
            extra = getattr(issue, "extra", None) or {}
            linking = extra.get("occurrence_urls") or []
            # Uncapped on purpose, and this time verifiably so: the PDF caps its
            # evidence list and points the reader here for the rest.
            try:
                evidence = evidence_for_excel(issue.issue_code, extra)
            except Exception:
                evidence = ""
            ws.append([
                issue.severity.upper(),
                issue.page_url or "Site-wide",
                issue.issue_code,
                issue.description,
                issue.recommendation,
                evidence,
                "\n".join(linking) if linking else "",
                extra.get("occurrence_urls_total") if linking else None,
            ])

        # Formatting
        ws.column_dimensions['A'].width = 12
        ws.column_dimensions['B'].width = 60
        ws.column_dimensions['C'].width = 25
        ws.column_dimensions['D'].width = 60
        ws.column_dimensions['E'].width = 60
        ws.column_dimensions['F'].width = 70
        ws.column_dimensions['G'].width = 70
        ws.column_dimensions['H'].width = 20
        
        # Add auto-filter
        ws.auto_filter.ref = ws.dimensions

    # ── Output ─────────────────────────────────────────────────────────────
    out = io.BytesIO()
    wb.save(out)
    return out.getvalue()
