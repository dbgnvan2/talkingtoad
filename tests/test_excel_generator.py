import io
import pytest
import openpyxl
from api.models.job import CrawlJob
from api.models.issue import Issue
from api.services.excel_generator import generate_excel_report


class TestConfidenceInExcel:
    """M7.b — confidence_label renders in Excel AI Readiness sheet."""

    def test_confidence_column_populated(self):
        """AI Readiness sheet has Confidence header and Established value."""
        job = CrawlJob(target_url="https://example.com")
        issue = Issue(
            job_id=job.job_id,
            page_url="https://example.com/page1",
            category="ai_readiness",
            severity="warning",
            issue_code="LLMS_TXT_MISSING",
            description="Missing llms.txt",
            recommendation="Add llms.txt",
            confidence_label="Established",
        )
        summary = {
            "health_score": 80,
            "pages_crawled": 1,
            "total_issues": 1,
            "by_severity": {"warning": 1},
            "by_category": {"ai_readiness": 1},
        }
        xlsx_bytes = generate_excel_report(job, [issue], summary)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["AI Readiness"]
        # Find header row with "Confidence"
        confidence_col = None
        for row in ws.iter_rows(min_row=1, max_row=30, max_col=10):
            for cell in row:
                if cell.value == "Confidence":
                    confidence_col = cell.column
                    break
            if confidence_col:
                break
        assert confidence_col is not None, "Confidence header not found"
        # Find data row with Established
        found = False
        for row in ws.iter_rows(
            min_row=2, max_row=ws.max_row,
            min_col=confidence_col, max_col=confidence_col,
        ):
            if row[0].value == "Established":
                found = True
                break
        assert found, "Established confidence value not found in sheet"

    def test_confidence_none_blank_cell(self):
        """Issue with confidence_label=None produces blank cell, no crash."""
        job = CrawlJob(target_url="https://example.com")
        issue = Issue(
            job_id=job.job_id,
            page_url="https://example.com/page2",
            category="ai_readiness",
            severity="info",
            issue_code="LLMS_TXT_MISSING",
            description="No llms.txt",
            recommendation="Add llms.txt",
            confidence_label=None,
        )
        summary = {
            "health_score": 90,
            "pages_crawled": 1,
            "total_issues": 1,
            "by_severity": {"info": 1},
            "by_category": {"ai_readiness": 1},
        }
        xlsx_bytes = generate_excel_report(job, [issue], summary)
        wb = openpyxl.load_workbook(io.BytesIO(xlsx_bytes))
        ws = wb["AI Readiness"]
        # Find the data row for this issue
        found_issue_row = False
        for row in ws.iter_rows(min_row=2, max_row=ws.max_row, max_col=10):
            if row[0].value == "LLMS_TXT_MISSING":
                confidence_cell = row[2]  # Column C = Confidence
                assert confidence_cell.value is None or confidence_cell.value == ""
                found_issue_row = True
                break
        assert found_issue_row, "Issue row not found in AI Readiness sheet"
