"""E7 — remediation roadmap and an honest scope statement.

Purpose: prove every roadmap row has an owner and a countable exit condition,
         that phases are derived rather than decorative, and — above all — that
         an omitted section is named in Caveats rather than reading as a pass.
Spec:    docs/pending/2026-08-29_E7-report-roadmap-and-caveats.md
Tests:   this file

The two honesty tests come first (P10). They are the guarantees: the places
where a silent omission would otherwise read to a client as a clean bill of
health, and E7.4b is the direct report-surface consequence of the E1 bug — a
"97% Image Health" printed over 13 of 1,284 images.
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from pathlib import Path

import pytest

from api.config import load_config
from api.crawler.checkers.registry import CATEGORY_DISPLAY, _CATALOGUE
from api.models.issue import Issue as IssueModel
from api.models.job import CrawlJob
from api.services.prevalence import compute_prevalence
from api.services.remediation import (
    build_roadmap,
    done_when_for,
    effort_label,
    owner_for,
    phase_titles,
)
from api.services.report_generator import generate_pdf_report

JOB_ID = "job-e7"


def _job(**kwargs) -> CrawlJob:
    return CrawlJob(
        job_id=JOB_ID, target_url="https://example.com",
        status="complete", started_at=datetime.now(timezone.utc), **kwargs
    )


def _issue(code: str, url: str = "https://example.com/p", *, rank: int = 10,
           impact: int = 4, effort: int = 1) -> IssueModel:
    spec = _CATALOGUE[code]
    return IssueModel(
        job_id=JOB_ID, page_url=url, category=spec.category, severity=spec.severity,
        issue_code=code, description="d", recommendation="r",
        impact=impact, effort=effort, priority_rank=rank,
        human_description=spec.human_description or code,
        fixability=spec.fixability,
    )


def _pdf_text(pdf_bytes: bytes) -> str:
    import pypdf

    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    return " ".join((p.extract_text() or "") for p in reader.pages).replace("\n", " ")


def _summary(**kwargs) -> dict:
    base = {"health_score": 90, "agent_health_score": 90, "pages_crawled": 10,
            "total_issues": 3, "by_severity": {"critical": 0, "warning": 1, "info": 2},
            "by_category": {}}
    base.update(kwargs)
    return base


# ── E7.4 — the honesty guarantees (written first) ──────────────────────────


class TestOmissionsAreDisclosed:
    @pytest.mark.asyncio
    async def test_e7_4a_omission_recorded_in_caveats(self):
        """No ledger → the section is absent AND named, so missing data can never
        be mistaken for a finding of no traffic."""
        pdf = await generate_pdf_report(_job(), [], _summary(), performance=None)
        text = _pdf_text(pdf)
        assert "Scope, Method and Caveats" in text
        assert "were omitted" in text
        assert "not a finding of no traffic" in text

    @pytest.mark.asyncio
    async def test_e7_4b_no_image_score_without_images(self):
        """The direct consequence of the E1 bug: no images collected must never
        produce a health percentage, and must say so."""
        pdf = await generate_pdf_report(_job(), [], _summary(), image_summary={"total_images": 0})
        text = _pdf_text(pdf)
        assert "Image Health Score" not in text
        assert "no images were collected" in text
        assert "not a statement that the site's images are fine" in text

    @pytest.mark.asyncio
    async def test_e7_2a_caveats_always_rendered_on_a_clean_site(self):
        """Zero issues is not a reason to hide what was not checked."""
        pdf = await generate_pdf_report(_job(), [], _summary(total_issues=0))
        assert "Scope, Method and Caveats" in _pdf_text(pdf)

    @pytest.mark.asyncio
    async def test_e7_2d_not_checked_list_is_complete(self):
        text = _pdf_text(await generate_pdf_report(_job(), [], _summary()))
        for expected in ("Off-site authority", "Core Web Vitals", "Server logs",
                         "CMS and plugin configuration", "WCAG conformance"):
            assert expected in text, f"Caveats must name {expected!r}"

    @pytest.mark.asyncio
    async def test_e7_2b_caps_disclosed_with_counts(self):
        job = _job(images_seen_total=1284, images_collected=150)
        text = _pdf_text(await generate_pdf_report(job, [], _summary()))
        assert "analysed 150 of 1284" in text.replace(",", "")

    @pytest.mark.asyncio
    async def test_e7_2c_unbitten_cap_is_silent(self):
        """A cap that did not bite must not be mentioned — noise costs trust."""
        job = _job(images_seen_total=12, images_collected=12)
        text = _pdf_text(await generate_pdf_report(job, [], _summary()))
        assert "Limits reached during this crawl" not in text

    @pytest.mark.asyncio
    async def test_e7_2e_score_meanings_stated(self):
        text = _pdf_text(await generate_pdf_report(_job(), [], _summary()))
        assert "What the scores mean" in text
        assert "does not forecast" in text or "None of them forecasts" in text


# ── E7.1 — the roadmap ─────────────────────────────────────────────────────


class TestRoadmap:
    def test_e7_1d_every_row_has_owner_and_done_when(self):
        issues = [_issue("META_DESC_MISSING"), _issue("H1_MISSING"),
                  _issue("BROKEN_LINK_404")]
        items, _weighted, _totals = build_roadmap(issues)
        assert items
        for item in items:
            assert item.owner and item.owner.strip()
            assert item.done_when and item.done_when.strip()

    def test_e7_1e_done_when_is_countable(self):
        """Every criterion must be checkable by re-running the tool."""
        countable = ("re-crawl", "decision")
        for code in ("META_DESC_MISSING", "BROKEN_LINK_404", "ENTITY_HOURS_DEFAULT",
                     "TITLE_TOO_LONG", "SEMANTIC_DENSITY_LOW"):
            text = done_when_for(code).casefold()
            assert any(w in text for w in countable), f"{code}: {text!r}"

    def test_e7_1e_fallback_is_countable_for_every_code(self):
        for code in _CATALOGUE:
            text = done_when_for(code).casefold()
            assert "re-crawl" in text or "decision" in text, code

    def test_e7_1f_owner_map_covers_all_categories(self):
        """No category may fall through to the default silently."""
        cfg = load_config("remediation_owners", required_keys=("by_category",))
        for key, _label in CATEGORY_DISPLAY:
            assert key in cfg["by_category"], f"no owner mapped for category {key}"

    def test_e7_1f_owner_lookup_falls_back_safely(self):
        assert owner_for("a_category_that_does_not_exist")

    def test_e7_1a_systemic_lands_in_phase_one(self):
        issues = [_issue("CONSENT_MODE_MISSING", f"https://example.com/p{i}")
                  for i in range(50)]
        prevalence = compute_prevalence(
            [("CONSENT_MODE_MISSING", f"https://example.com/p{i}") for i in range(50)], 60
        )
        items, weighted, _totals = build_roadmap(issues, prevalence=prevalence)
        assert weighted is True
        assert items[0].phase == "phase_1"
        assert items[0].tier == "systemic"

    def test_e7_1b_high_traffic_page_lands_in_phase_two(self):
        issues = [_issue("TITLE_TOO_LONG", "https://example.com/earner")]
        priority = [{"url": "https://example.com/earner", "priority_rank": 1},
                    {"url": "https://example.com/other", "priority_rank": 2},
                    {"url": "https://example.com/x", "priority_rank": 3},
                    {"url": "https://example.com/y", "priority_rank": 4}]
        items, _weighted, _totals = build_roadmap(issues, priority_pages=priority)
        assert items[0].phase == "phase_2"

    def test_e7_1c_graceful_without_priority_data(self):
        items, weighted, _totals = build_roadmap([_issue("TITLE_TOO_LONG")])
        assert weighted is False
        assert all(i.phase == "phase_3" for i in items)

    def test_e7_1a_deduplicates_by_code(self):
        issues = [_issue("META_DESC_MISSING", f"https://example.com/p{i}") for i in range(20)]
        items, _weighted, _totals = build_roadmap(issues)
        assert len([i for i in items if i.code == "META_DESC_MISSING"]) == 1

    def test_e7_1a_pages_affected_counted(self):
        issues = [_issue("META_DESC_MISSING", f"https://example.com/p{i}") for i in range(7)]
        items, _weighted, _totals = build_roadmap(issues)
        assert items[0].pages_affected == 7

    def test_e7_1a_per_phase_cap_applied(self):
        issues = [_issue(code) for code in list(_CATALOGUE)[:40]]
        items, _weighted, totals = build_roadmap(issues, limit_per_phase=5)
        for _name, _title in phase_titles():
            assert len([i for i in items if i.phase_title == _title]) <= 5

    def test_e7_1a_cap_returns_pre_cap_totals_so_it_can_be_disclosed(self):
        """Rule 6. Returning only the capped list made disclosure impossible —
        no caller could know 28 of 40 items had been dropped."""
        issues = [_issue(code) for code in list(_CATALOGUE)[:40]]
        items, _weighted, totals = build_roadmap(issues, limit_per_phase=5)
        assert sum(totals.values()) == 40
        assert sum(totals.values()) > len(items)

    def test_e7_1a_uncapped_call_returns_everything(self):
        """The Excel Roadmap sheet is where the PDF points for the full list, so
        an uncapped call must genuinely return it (F2/F6)."""
        issues = [_issue(code) for code in list(_CATALOGUE)[:40]]
        items, _weighted, totals = build_roadmap(issues, limit_per_phase=10**6)
        assert len(items) == sum(totals.values()) == 40

    def test_e7_1a_empty_input_is_not_a_crash(self):
        assert build_roadmap([]) == ([], False, {})

    @pytest.mark.parametrize("effort,expected", [(0, "Trivial"), (1, "Low"),
                                                 (2, "Medium"), (3, "High")])
    def test_e7_1a_effort_labels(self, effort, expected):
        assert effort_label(effort) == expected


class TestRoadmapRendered:
    @pytest.mark.asyncio
    async def test_e7_1a_roadmap_section_present(self):
        issues = [_issue("META_DESC_MISSING", f"https://example.com/p{i}") for i in range(30)]
        prevalence = compute_prevalence(
            [("META_DESC_MISSING", f"https://example.com/p{i}") for i in range(30)], 40
        )
        pdf = await generate_pdf_report(_job(), issues, _summary(), prevalence=prevalence)
        text = _pdf_text(pdf)
        assert "Remediation Roadmap" in text
        assert "Owner:" in text
        assert "Done when:" in text
        assert "Phase 1" in text

    @pytest.mark.asyncio
    async def test_e7_1c_roadmap_says_so_when_unweighted(self):
        """No prevalence and no traffic data → the section must not imply phasing
        it did not apply."""
        pdf = await generate_pdf_report(_job(), [_issue("TITLE_TOO_LONG")], _summary())
        text = _pdf_text(pdf)
        assert "Remediation Roadmap" in text
        assert "No prevalence or traffic data was available" in text

    @pytest.mark.asyncio
    async def test_e7_1a_no_issues_no_roadmap(self):
        pdf = await generate_pdf_report(_job(), [], _summary(total_issues=0))
        assert "Remediation Roadmap" not in _pdf_text(pdf)


class TestExcelRoadmap:
    """E7 surface parity (P25) — the roadmap must reach the spreadsheet too."""

    def test_e7_excel_roadmap_tab(self):
        from openpyxl import load_workbook

        from api.services.excel_generator import generate_excel_report

        issues = [_issue("META_DESC_MISSING", f"https://example.com/p{i}") for i in range(5)]
        xlsx = generate_excel_report(_job(), issues, _summary())
        wb = load_workbook(io.BytesIO(xlsx))
        assert "Roadmap" in wb.sheetnames
        headers = [c.value for c in wb["Roadmap"][3]]
        assert "Owner" in headers and "Done when" in headers

    def test_e7_excel_roadmap_omitted_without_issues(self):
        from openpyxl import load_workbook

        from api.services.excel_generator import generate_excel_report

        xlsx = generate_excel_report(_job(), [], _summary(total_issues=0))
        assert "Roadmap" not in load_workbook(io.BytesIO(xlsx)).sheetnames


class TestReviewFixes:
    """Regressions found by the pre-push failure-pattern review (2026-08-29).

    Each of these was a real defect in the E-series work — mostly a claim in
    report or UI copy that the data behind it did not support.
    """

    @pytest.mark.asyncio
    async def test_f8_tooling_failure_is_not_reported_as_missing_client_data(self):
        """`performance=None` has two causes: the client supplied no GSC/GA4
        data, or our own guard fired. Printing "no data was supplied for this
        site" in the second case is a false statement about their inputs."""
        pdf = await generate_pdf_report(
            _job(), [], _summary(), performance=None, performance_failed=True
        )
        text = _pdf_text(pdf)
        assert "could not be read while building this report" in text
        assert "tooling failure on our side" in text
        assert "was supplied for this site" not in text

    @pytest.mark.asyncio
    async def test_f8_genuinely_absent_data_keeps_the_original_wording(self):
        pdf = await generate_pdf_report(
            _job(), [], _summary(), performance=None, performance_failed=False
        )
        text = _pdf_text(pdf)
        assert "no Search Console" in text
        assert "tooling failure" not in text

    @pytest.mark.asyncio
    async def test_f6_roadmap_cap_is_disclosed(self):
        """Rule 6 — a phase showing 12 of 40 must say so."""
        issues = [_issue(code) for code in list(_CATALOGUE)[:40]]
        pdf = await generate_pdf_report(_job(), issues, _summary())
        text = _pdf_text(pdf)
        assert "Showing the top" in text and " of " in text
        assert "Excel Roadmap sheet" in text

    @pytest.mark.asyncio
    async def test_f6_no_cap_disclosure_when_nothing_dropped(self):
        pdf = await generate_pdf_report(_job(), [_issue("TITLE_TOO_LONG")], _summary())
        assert "Showing the top" not in _pdf_text(pdf)

    @pytest.mark.asyncio
    async def test_f5_systemic_heading_does_not_claim_a_share_it_cannot_keep(self):
        """`always_systemic` codes bypass the share gate, so a 1%-footprint code
        can be systemic. The heading must not promise "30% or more"."""
        from api.services.prevalence import compute_prevalence

        prevalence = compute_prevalence(
            [("BROKEN_LINK_404", f"https://x/p{i}") for i in range(3)], 272
        )
        assert prevalence[0].tier == "systemic" and prevalence[0].share < 0.05
        pdf = await generate_pdf_report(_job(), [], _summary(), prevalence=prevalence)
        text = _pdf_text(pdf)
        assert "Systemic Defects" in text
        # The copy may cite a large share as ONE of two reasons, but must not
        # assert it unconditionally — this code qualifies on the other one.
        assert "or because the fix is inherently a template" in text
        assert "3 of 272" in text, "the real footprint must be stated"

    @pytest.mark.asyncio
    async def test_f11_hygiene_note_matches_the_formula_in_use(self):
        from api.services.prevalence import compute_prevalence

        prevalence = compute_prevalence(
            [("CONSENT_MODE_MISSING", f"https://x/p{i}") for i in range(50)], 60
        )
        text = _pdf_text(await generate_pdf_report(_job(), [], _summary(), prevalence=prevalence))
        assert "carrying NO systemic defect" in text
        assert "weighted by how widespread" not in text, "stale formula text"
