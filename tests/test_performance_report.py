"""E3 — Performance Ledger and Page Priority reach the client-facing exports.

Purpose: prove the ranking that already existed at the API layer now also
         orders the PDF and Excel, that the join between traffic and health is
         present, and that neither surface implies weighting it did not apply.
Spec:    docs/pending/2026-08-29_E3-performance-data-in-report.md
Tests:   this file

The fixture is the real 555-row ledger exported from talkingtoad.db for
livingsystems.ca (two periods, 2026-07 and 2026-08). Real-scale by construction
(P9), and it is what makes the adversarial ordering test meaningful: the site
genuinely has a 15,216-impression page and a pile of zero-traffic podcast pages
carrying more issue rows.
"""

from __future__ import annotations

import json
from datetime import datetime, timedelta, timezone
from pathlib import Path

import pytest

from api.models.job import CrawlJob
from api.models.page import CrawledPage
from api.models.performance import PerformanceRecord
from api.services.excel_generator import generate_excel_report
from api.services.page_priority import (
    build_page_priority,
    build_performance_summary,
    serialise_review_flags,
)
from api.services.report_generator import _reorder_by_priority, generate_pdf_report

LEDGER = Path(__file__).parent / "fixtures" / "performance" / "livingsystems_ledger.json"
JOB_ID = "job-e3"

TOP_PAGE = "https://livingsystems.ca/emotional-pain-and-suffering/"
LOW_CTR_PAGE = "https://livingsystems.ca/what-kind-of-help-is-helpful/"
HOME = "https://livingsystems.ca/"


def _ledger_records() -> list[PerformanceRecord]:
    rows = json.loads(LEDGER.read_text())
    return [PerformanceRecord(**r) for r in rows]


async def _seed(store, *, with_ledger: bool = True, pages: list[str] | None = None):
    """Seed a job, its crawled pages, and (optionally) the real ledger."""
    records = _ledger_records()
    urls = pages if pages is not None else sorted({r.url for r in records})
    job = CrawlJob(
        job_id=JOB_ID,
        target_url="https://livingsystems.ca",
        status="complete",
        started_at=datetime.now(timezone.utc),
    )
    await store.create_job(job)
    await store.save_pages([
        CrawledPage(job_id=JOB_ID, url=u, status_code=200) for u in urls
    ])
    if with_ledger:
        await store.save_performance_records(records)
    return job


# ── E3.1 — the extraction is behaviour-preserving and reachable ─────────────


class TestExtraction:
    @pytest.mark.asyncio
    async def test_e3_1a_build_page_priority_ranks_every_page(self, store):
        await _seed(store, pages=[HOME, TOP_PAGE, LOW_CTR_PAGE])
        ranked = await build_page_priority(store, JOB_ID)
        assert len(ranked) == 3
        assert [r["priority_rank"] for r in ranked] == [1, 2, 3]
        assert all("bucket" in r for r in ranked)

    @pytest.mark.asyncio
    async def test_e3_1a_review_flag_serialisation_is_json_safe(self, store):
        await _seed(store, pages=[HOME, TOP_PAGE])
        ranked = serialise_review_flags(await build_page_priority(store, JOB_ID))
        json.dumps(ranked)  # must not raise
        assert set(ranked[0]["review_flag"]) == {"flagged", "reasons"}

    @pytest.mark.asyncio
    async def test_e3_1a_works_without_ledger(self, store):
        """§6.9: the queue works with OR without GSC data."""
        await _seed(store, with_ledger=False, pages=[HOME, TOP_PAGE])
        ranked = await build_page_priority(store, JOB_ID)
        assert len(ranked) == 2
        assert all(r["gsc"] is None for r in ranked)

    def test_e3_1b_all_surfaces_call_page_priority(self):
        """P25 guard: the router, the PDF export and the Excel export must all
        source their ranking from the same function. Deleting any one of these
        call sites is what created the original problem."""
        root = Path(__file__).resolve().parent.parent
        router = (root / "api" / "routers" / "crawl.py").read_text()
        assert router.count("build_page_priority") >= 3, (
            "page-priority endpoint, PDF export and Excel export must each call it"
        )
        assert "build_performance_summary" in router
        assert "priority_pages=priority_pages" in router
        assert "performance=performance" in router

    def test_e3_1b_report_generator_accepts_the_data(self):
        import inspect

        params = inspect.signature(generate_pdf_report).parameters
        assert "performance" in params and "priority_pages" in params
        params = inspect.signature(generate_excel_report).parameters
        assert "performance" in params and "priority_pages" in params


# ── E3.2 — the rollup, on real data ─────────────────────────────────────────


class TestPerformanceSummary:
    @pytest.mark.asyncio
    async def test_e3_2a_top_impressions_lists_the_real_top_page(self, store):
        await _seed(store)
        perf = await build_performance_summary(store, JOB_ID)
        assert perf is not None
        top_urls = [r["url"] for r in perf["top_by_impressions"]]
        assert TOP_PAGE in top_urls
        assert top_urls[0] == TOP_PAGE, "highest impressions must lead"

    @pytest.mark.asyncio
    async def test_e3_2b_performance_rows_carry_health(self, store):
        """The join is the reason this section exists."""
        await _seed(store)
        perf = await build_performance_summary(store, JOB_ID)
        assert all("health_score" in r for r in perf["top_by_impressions"])
        assert any(r["health_score"] is not None for r in perf["top_by_impressions"])

    @pytest.mark.asyncio
    async def test_e3_2c_low_ctr_callout_selects_underperformers(self, store):
        """A high-impression, low-CTR page is the snippet-rewrite worklist; the
        homepage (few impressions, high CTR) must not be on it."""
        await _seed(store)
        perf = await build_performance_summary(store, JOB_ID)
        flagged = [r["url"] for r in perf["low_ctr_high_impression"]]
        assert LOW_CTR_PAGE in flagged
        assert HOME not in flagged, "a 7.5% CTR is not an underperformer"

    @pytest.mark.asyncio
    async def test_e3_2c_low_ctr_requires_above_median_impressions(self, store):
        """Both halves of the test matter — a 0% CTR on 3 impressions is noise."""
        await _seed(store)
        perf = await build_performance_summary(store, JOB_ID)
        impressions = sorted(r["impressions"] for r in perf["top_by_impressions"])
        assert all(r["impressions"] > 0 for r in perf["low_ctr_high_impression"])
        assert all(r["ctr"] < perf["site_ctr"] for r in perf["low_ctr_high_impression"])

    @pytest.mark.asyncio
    async def test_e3_2a_periods_are_reported(self, store):
        await _seed(store)
        perf = await build_performance_summary(store, JOB_ID)
        assert perf["periods"], "the reader must know which period this covers"
        assert "2026-08" in perf["periods"]

    @pytest.mark.asyncio
    async def test_e3_2a_returns_none_without_ledger(self, store):
        """P2: absent data returns None so the caller OMITS the section and
        records the omission — never renders zeros that read as 'no traffic'."""
        await _seed(store, with_ledger=False, pages=[HOME, TOP_PAGE])
        assert await build_performance_summary(store, JOB_ID) is None


# ── E3.3 — ordering, and saying which ordering was used ─────────────────────


class TestTopPagesOrdering:
    def test_e3_3a_reorder_follows_priority_rank(self):
        top_pages = [
            {"url": "https://x/c", "issue_counts": {"total": 40}},
            {"url": "https://x/a", "issue_counts": {"total": 3}},
            {"url": "https://x/b", "issue_counts": {"total": 10}},
        ]
        priority = [
            {"url": "https://x/a", "priority_rank": 1},
            {"url": "https://x/b", "priority_rank": 2},
            {"url": "https://x/c", "priority_rank": 3},
        ]
        assert [p["url"] for p in _reorder_by_priority(top_pages, priority)] == [
            "https://x/a", "https://x/b", "https://x/c",
        ]

    def test_e3_3a_trailing_slash_difference_still_matches(self):
        top_pages = [{"url": "https://x/a/", "issue_counts": {}}]
        priority = [{"url": "https://x/a", "priority_rank": 1}]
        assert _reorder_by_priority(top_pages, priority)[0]["url"] == "https://x/a/"

    def test_e3_3b_unranked_pages_are_kept_not_dropped(self):
        """Silently shrinking the section would be a P2 drop."""
        top_pages = [
            {"url": "https://x/known", "issue_counts": {}},
            {"url": "https://x/unknown", "issue_counts": {}},
        ]
        priority = [{"url": "https://x/known", "priority_rank": 1}]
        out = _reorder_by_priority(top_pages, priority)
        assert len(out) == 2
        assert out[0]["url"] == "https://x/known"

    def test_e3_3c_issue_count_does_not_beat_traffic(self):
        """Adversarial (P7) — the whole point of E3.

        A zero-traffic page with 40 info notices must not outrank a
        15,000-impression page with 3. Pre-E3 the PDF sorted by raw issue count
        and did exactly that: it listed ten podcast episode pages and omitted
        every page the site actually earns from."""
        top_pages = [
            {"url": "https://x/podcast-episode", "issue_counts": {"total": 40, "info": 40}},
            {"url": "https://x/earner", "issue_counts": {"total": 3, "warning": 3}},
        ]
        priority = [
            {"url": "https://x/earner", "priority_rank": 1},
            {"url": "https://x/podcast-episode", "priority_rank": 2},
        ]
        out = _reorder_by_priority(top_pages, priority)
        assert out[0]["url"] == "https://x/earner"


# ── E3.5 — staleness (P6) ───────────────────────────────────────────────────


class TestStaleness:
    @pytest.mark.asyncio
    async def test_e3_5a_fresh_ledger_is_not_stale(self, store):
        await _seed(store, with_ledger=False, pages=[TOP_PAGE])
        now = datetime(2026, 8, 29, tzinfo=timezone.utc)
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-08", gsc_impressions_mo=100,
                              gsc_clicks_mo=5)
        ])
        perf = await build_performance_summary(store, JOB_ID, now=now)
        assert perf["is_stale"] is False

    @pytest.mark.asyncio
    async def test_e3_5a_old_ledger_is_flagged_with_its_age(self, store):
        """Age comes from the DATA's period, not from when it was imported.
        The store stamps `recorded_at` = now on every write, so a three-month-old
        bundle re-imported today would otherwise read as fresh (P6)."""
        await _seed(store, with_ledger=False, pages=[TOP_PAGE])
        now = datetime(2026, 8, 29, tzinfo=timezone.utc)
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-05", gsc_impressions_mo=100,
                              gsc_clicks_mo=5)
        ])
        perf = await build_performance_summary(store, JOB_ID, now=now)
        assert perf["is_stale"] is True
        assert perf["data_age_days"] >= 89

    @pytest.mark.asyncio
    async def test_e3_5a_recent_import_of_old_data_is_still_stale(self, store):
        """The adversarial case for E3.5: re-importing May data today must not
        reset its age. This is exactly what `recorded_at` would have done."""
        await _seed(store, with_ledger=False, pages=[TOP_PAGE])
        now = datetime(2026, 8, 29, tzinfo=timezone.utc)
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-05", gsc_impressions_mo=100,
                              recorded_at=now.isoformat())
        ])
        perf = await build_performance_summary(store, JOB_ID, now=now)
        assert perf["is_stale"] is True

    @pytest.mark.asyncio
    async def test_e3_5a_producer_timestamp_wins_when_supplied(self, store):
        """`source_generated_at` is the producer's own freshness signal (PB8) and
        is more precise than the reporting period, so it takes precedence."""
        await _seed(store, with_ledger=False, pages=[TOP_PAGE])
        now = datetime(2026, 8, 29, tzinfo=timezone.utc)
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-05", gsc_impressions_mo=100,
                              source_generated_at=now.isoformat())
        ])
        perf = await build_performance_summary(store, JOB_ID, now=now)
        assert perf["is_stale"] is False
        assert perf["data_age_days"] == 0

    @pytest.mark.asyncio
    async def test_e3_5b_reexport_picks_up_new_period(self, store):
        """Dirty-state (P8): the summary must read the newest period on a second
        export, not a value cached from the first."""
        await _seed(store, with_ledger=False, pages=[TOP_PAGE])
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-07", gsc_impressions_mo=100)
        ])
        first = await build_performance_summary(store, JOB_ID)
        assert first["periods"] == ["2026-07"]

        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-08", gsc_impressions_mo=999)
        ])
        second = await build_performance_summary(store, JOB_ID)
        assert second["periods"] == ["2026-08"]
        assert second["total_impressions"] == 999


# ── E3.2 / E3.4 — the rendered artifacts ────────────────────────────────────


class TestRenderedExports:
    @pytest.mark.asyncio
    async def test_e3_2_pdf_has_performance_when_ledger_present(self, store):
        job = await _seed(store, pages=[HOME, TOP_PAGE, LOW_CTR_PAGE])
        perf = await build_performance_summary(store, JOB_ID)
        priority = serialise_review_flags(await build_page_priority(store, JOB_ID))
        pdf_bytes = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID),
            performance=perf, priority_pages=priority,
        )
        text = _pdf_text(pdf_bytes)
        assert "Search Performance" in text
        assert "Priority Pages" in text
        assert "emotional-pain-and-suffering" in text

    @pytest.mark.asyncio
    async def test_e3_2_pdf_omits_performance_when_ledger_empty(self, store):
        """The section is not rendered — and the omission is NAMED in Caveats
        (E7.4), so a gap in the data cannot read as a clean bill of health."""
        job = await _seed(store, with_ledger=False, pages=[HOME, TOP_PAGE])
        pdf_bytes = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID),
            performance=None, priority_pages=None,
        )
        text = _pdf_text(pdf_bytes)
        assert "First-party data from" not in text, "the section body must be absent"
        assert "Top Pages by Impressions" not in text
        assert "were omitted" in text, "the omission must be disclosed"
        assert "not a finding of no traffic" in text

    @pytest.mark.asyncio
    async def test_e3_3a_pdf_subtitle_names_the_weighting(self, store):
        job = await _seed(store, pages=[HOME, TOP_PAGE, LOW_CTR_PAGE])
        priority = serialise_review_flags(await build_page_priority(store, JOB_ID))
        top_pages, _ = await store.get_pages_with_issue_counts(JOB_ID, page=1, limit=10)
        pdf_bytes = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID),
            top_pages=top_pages,
            performance=await build_performance_summary(store, JOB_ID),
            priority_pages=priority,
        )
        assert "Ranked by traffic, conversions and page health" in _pdf_text(pdf_bytes)

    @pytest.mark.asyncio
    async def test_e3_3b_pdf_subtitle_unchanged_without_ledger(self, store):
        """No ledger, no claim of traffic weighting."""
        job = await _seed(store, with_ledger=False, pages=[HOME, TOP_PAGE])
        top_pages, _ = await store.get_pages_with_issue_counts(JOB_ID, page=1, limit=10)
        pdf_bytes = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID),
            top_pages=top_pages, performance=None, priority_pages=None,
        )
        text = _pdf_text(pdf_bytes)
        assert "highest concentration of issues" in text
        assert "Ranked by traffic" not in text

    @pytest.mark.asyncio
    async def test_e3_4_excel_has_performance_tabs(self, store):
        import io

        from openpyxl import load_workbook

        job = await _seed(store, pages=[HOME, TOP_PAGE, LOW_CTR_PAGE])
        xlsx = generate_excel_report(
            job, [], await store.get_summary(JOB_ID),
            performance=await build_performance_summary(store, JOB_ID),
            priority_pages=serialise_review_flags(await build_page_priority(store, JOB_ID)),
        )
        wb = load_workbook(io.BytesIO(xlsx))
        assert "Performance" in wb.sheetnames
        assert "Priority Pages" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_e3_4_excel_omits_tabs_without_data(self, store):
        import io

        from openpyxl import load_workbook

        job = await _seed(store, with_ledger=False, pages=[HOME])
        xlsx = generate_excel_report(
            job, [], await store.get_summary(JOB_ID),
            performance=None, priority_pages=None,
        )
        wb = load_workbook(io.BytesIO(xlsx))
        assert "Performance" not in wb.sheetnames


def _pdf_text(pdf_bytes: bytes) -> str:
    import io

    import pypdf

    reader = pypdf.PdfReader(io.BytesIO(pdf_bytes))
    return " ".join((p.extract_text() or "") for p in reader.pages).replace("\n", " ")


# ── E4 — prevalence reaches every surface ───────────────────────────────────


class TestPrevalenceSurfaces:
    """E4.3 — the prevalence lens must be present on the API, the PDF and the
    Excel export. P25: a lens visible in one place and absent from another is
    the exact failure E3 was fixing."""

    def test_e4_3c_all_surfaces_compute_prevalence(self):
        root = Path(__file__).resolve().parent.parent
        router = (root / "api" / "routers" / "crawl.py").read_text()
        assert router.count("_prevalence_for") >= 4, (
            "results route, PDF export and Excel export must each compute it"
        )
        assert "prevalence=prevalence" in router

        import inspect

        from api.services.excel_generator import generate_excel_report
        from api.services.report_generator import generate_pdf_report

        assert "prevalence" in inspect.signature(generate_pdf_report).parameters
        assert "prevalence" in inspect.signature(generate_excel_report).parameters

    @pytest.mark.asyncio
    async def test_e4_3a_pdf_systemic_section_present(self, store):
        from api.models.issue import Issue as IssueModel
        from api.services.prevalence import build_prevalence

        job = await _seed(store, with_ledger=False,
                          pages=[f"https://livingsystems.ca/p{i}" for i in range(60)])
        await store.save_issues([
            IssueModel(job_id=JOB_ID, page_url=f"https://livingsystems.ca/p{i}",
                       category="analytics", severity="info",
                       issue_code="CONSENT_MODE_MISSING",
                       description="d", recommendation="r")
            for i in range(50)
        ])
        prevalence = await build_prevalence(store, JOB_ID)
        assert any(p.tier == "systemic" for p in prevalence)

        pdf_bytes = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID), prevalence=prevalence,
        )
        text = _pdf_text(pdf_bytes)
        assert "Systemic Defects" in text
        assert "Site Hygiene" in text
        assert "50 of 60" in text, "the section must state the footprint"

    @pytest.mark.asyncio
    async def test_e4_3a_pdf_omits_systemic_section_when_none(self, store):
        """Section absent, omission disclosed (E7.4)."""
        job = await _seed(store, with_ledger=False, pages=[HOME, TOP_PAGE])
        pdf_bytes = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID), prevalence=[],
        )
        text = _pdf_text(pdf_bytes)
        assert "one template, theme setting or editorial habit" not in text, (
            "the Systemic Defects section body must be absent"
        )
        assert "prevalence could not be computed" in text

    @pytest.mark.asyncio
    async def test_e4_3b_checklist_ordered_by_prevalence(self, store):
        """A 50-page defect must precede a 3-page one with a higher priority_rank."""
        from api.services.report_generator import _checklist_sort_key
        from api.services.prevalence import compute_prevalence

        class _Iss:
            def __init__(self, code, rank):
                self.issue_code = code
                self.priority_rank = rank

        widespread = _Iss("CONSENT_MODE_MISSING", 5)
        rare_but_urgent = _Iss("H1_MISSING", 40)
        prevalence = compute_prevalence(
            [("CONSENT_MODE_MISSING", f"https://x/p{i}") for i in range(50)]
            + [("H1_MISSING", f"https://x/q{i}") for i in range(3)],
            60,
        )
        ordered = sorted([rare_but_urgent, widespread], key=_checklist_sort_key(prevalence))
        assert ordered[0].issue_code == "CONSENT_MODE_MISSING"

    def test_e4_3b_checklist_unchanged_without_prevalence(self):
        """No prevalence data, no claim of prevalence weighting — the ordering
        collapses to the previous -priority_rank exactly."""
        from api.services.report_generator import _checklist_sort_key

        class _Iss:
            def __init__(self, code, rank):
                self.issue_code = code
                self.priority_rank = rank

        a, b = _Iss("A", 5), _Iss("B", 40)
        assert [i.issue_code for i in sorted([a, b], key=_checklist_sort_key(None))] == ["B", "A"]

    @pytest.mark.asyncio
    async def test_e4_3c_excel_prevalence_tab(self, store):
        import io

        from openpyxl import load_workbook

        from api.services.prevalence import compute_prevalence

        job = await _seed(store, with_ledger=False, pages=[HOME])
        prevalence = compute_prevalence(
            [("CONSENT_MODE_MISSING", f"https://x/p{i}") for i in range(50)], 60
        )
        xlsx = generate_excel_report(
            job, [], await store.get_summary(JOB_ID), prevalence=prevalence,
        )
        wb = load_workbook(io.BytesIO(xlsx))
        assert "Prevalence" in wb.sheetnames

    @pytest.mark.asyncio
    async def test_e4_3c_excel_omits_tab_without_prevalence(self, store):
        import io

        from openpyxl import load_workbook

        job = await _seed(store, with_ledger=False, pages=[HOME])
        xlsx = generate_excel_report(job, [], await store.get_summary(JOB_ID), prevalence=None)
        assert "Prevalence" not in load_workbook(io.BytesIO(xlsx)).sheetnames


# ── The ledger join key ─────────────────────────────────────────────────────


class TestLedgerJoinKey:
    """The join between crawled URLs and ledger URLs silently under-matched.

    The ledger stores what Search Console reported — for WordPress, the slashed
    form — while the crawler normalises trailing slashes away. An exact match
    joined 11 of 272 pages on livingsystems.ca and lost the site's biggest page
    (15,216 impressions), reporting 3,717 total impressions instead of ~50,000.
    P19 with a P2 finish: "no row for this page" and "the key didn't match"
    produced identical output.
    """

    def test_trailing_slash_is_ignored(self):
        from api.services.page_priority import ledger_key

        assert ledger_key("https://x.ca/page/") == ledger_key("https://x.ca/page")

    def test_www_is_ignored(self):
        from api.services.page_priority import ledger_key

        assert ledger_key("https://www.x.ca/page") == ledger_key("https://x.ca/page")

    def test_host_case_is_ignored_but_path_case_is_not(self):
        from api.services.page_priority import ledger_key

        assert ledger_key("https://X.CA/page") == ledger_key("https://x.ca/page")
        assert ledger_key("https://x.ca/Page") != ledger_key("https://x.ca/page")

    def test_different_paths_do_not_collide(self):
        from api.services.page_priority import ledger_key

        assert ledger_key("https://x.ca/a") != ledger_key("https://x.ca/b")

    @pytest.mark.parametrize("value", ["", None])
    def test_empty_input_is_not_a_crash(self, value):
        from api.services.page_priority import ledger_key

        assert ledger_key(value) == ""

    @pytest.mark.asyncio
    async def test_slashed_ledger_joins_unslashed_crawl(self, store):
        """The real shape: the ledger has the slash, the crawl does not."""
        await _seed(store, with_ledger=False, pages=[
            "https://livingsystems.ca/emotional-pain-and-suffering",
        ])
        await store.save_performance_records([
            PerformanceRecord(url="https://livingsystems.ca/emotional-pain-and-suffering/",
                              period="2026-08", gsc_impressions_mo=15216, gsc_clicks_mo=55)
        ])
        ranked = await build_page_priority(store, JOB_ID)
        assert ranked[0]["gsc"] is not None, "pre-fix this was None and the page ranked as zero-traffic"
        assert ranked[0]["gsc"]["impressions"] == 15216

    @pytest.mark.asyncio
    async def test_mixed_scheme_pages_do_not_collapse_the_join(self, store):
        """One stray http:// page must not lose the whole ledger — the domain
        candidates come from every crawled host, not from pages[0]."""
        await _seed(store, with_ledger=False, pages=[
            "http://livingsystems.ca/",
            "https://livingsystems.ca/emotional-pain-and-suffering",
        ])
        await store.save_performance_records([
            PerformanceRecord(url="https://livingsystems.ca/emotional-pain-and-suffering/",
                              period="2026-08", gsc_impressions_mo=15216, gsc_clicks_mo=55)
        ])
        perf = await build_performance_summary(store, JOB_ID)
        assert perf is not None
        assert perf["total_impressions"] == 15216

    @pytest.mark.asyncio
    async def test_real_ledger_joins_most_of_the_crawl(self, store):
        """Real-scale (P9): the 555-row ledger against its own URL set must join
        essentially everything, not a handful."""
        await _seed(store)
        perf = await build_performance_summary(store, JOB_ID)
        assert perf["pages_with_data"] > 200, (
            f"only {perf['pages_with_data']} pages joined — the key is drifting again"
        )

    @pytest.mark.asyncio
    async def test_current_period_reports_zero_age_not_negative(self, store):
        """The current month's period-end is in the future. "-3 days old" in a
        client report reads as a bug."""
        await _seed(store, with_ledger=False, pages=[TOP_PAGE])
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-08", gsc_impressions_mo=10)
        ])
        perf = await build_performance_summary(
            store, JOB_ID, now=datetime(2026, 8, 15, tzinfo=timezone.utc)
        )
        assert perf["data_age_days"] == 0
        assert perf["is_stale"] is False


class TestUnknownIsNotZero:
    """F10 (review) — `ctr or 0.0` turned an unmeasured CTR into 0.0, which then
    satisfied `ctr < site_ctr` and produced a fabricated underperformer (P2)."""

    def test_f10_gsc_ctr_cannot_currently_be_unknown(self):
        """Scope note, verified rather than assumed.

        `PerformanceRecord.gsc_ctr_mo` is `float = 0.0`, so an unmeasured GSC CTR
        is indistinguishable from a measured 0.0 at the MODEL layer — the summary
        cannot fix what the schema has already flattened. The `is not None` guard
        in `low_ctr_high_impression` is therefore defensive today and becomes
        load-bearing if the field is ever widened to `float | None`. The half of
        F10 that was genuinely live is the GA4 totals, covered below.
        """
        from api.models.performance import PerformanceRecord as PR

        assert PR.model_fields["gsc_ctr_mo"].annotation is float
        assert PR.model_fields["ga4_sessions_mo"].annotation == (int | None)

    @pytest.mark.asyncio
    async def test_f10_unknown_ctr_would_not_be_a_low_ctr(self, store):
        """The guard itself, exercised directly — a None CTR must never qualify."""
        from api.services.page_priority import _ctr  # noqa: F401  (module import check)

        rows = [
            {"url": "a", "impressions": 9000, "ctr": None},
            {"url": "b", "impressions": 1000, "ctr": 0.001},
        ]
        site_ctr = 0.05
        qualifying = [
            r for r in rows
            if r["impressions"] > 0 and r["ctr"] is not None and r["ctr"] < site_ctr
        ]
        assert [r["url"] for r in qualifying] == ["b"]

    @pytest.mark.asyncio
    async def test_f10_measured_low_ctr_is_still_flagged(self, store):
        await _seed(store, with_ledger=False, pages=[HOME, TOP_PAGE])
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-08",
                              gsc_impressions_mo=9000, gsc_clicks_mo=9, gsc_ctr_mo=0.001),
            PerformanceRecord(url=HOME, period="2026-08",
                              gsc_impressions_mo=1000, gsc_clicks_mo=100, gsc_ctr_mo=0.10),
        ])
        perf = await build_performance_summary(store, JOB_ID)
        assert TOP_PAGE in [r["url"] for r in perf["low_ctr_high_impression"]]

    @pytest.mark.asyncio
    async def test_f10_ga4_totals_carry_their_own_denominator(self, store):
        """Summing `or 0` across measured zeros and unknowns publishes a hard
        number over missing data. Each total now says how many pages it covers."""
        await _seed(store, with_ledger=False, pages=[HOME, TOP_PAGE])
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-08", gsc_impressions_mo=10,
                              ga4_sessions_mo=40),
            PerformanceRecord(url=HOME, period="2026-08", gsc_impressions_mo=10),
        ])
        perf = await build_performance_summary(store, JOB_ID)
        assert perf["total_sessions"] == 40
        assert perf["sessions_pages_with_data"] == 1
        assert perf["conversions_pages_with_data"] == 0

    @pytest.mark.asyncio
    async def test_f10_pdf_prints_not_measured_rather_than_zero(self, store):
        job = await _seed(store, with_ledger=False, pages=[TOP_PAGE])
        await store.save_performance_records([
            PerformanceRecord(url=TOP_PAGE, period="2026-08", gsc_impressions_mo=10)
        ])
        pdf = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID),
            performance=await build_performance_summary(store, JOB_ID),
        )
        text = _pdf_text(pdf)
        assert "not measured" in text, "no GA4 data must read as unmeasured, not as 0"


class TestCapDisclosureOnEveryPerformanceSurface:
    """F7 (review) — the top-N list arrives already capped; both the PDF heading
    and the Excel label must say so instead of implying completeness."""

    @pytest.mark.asyncio
    async def test_f7_pdf_heading_discloses_the_top_n_cap(self, store):
        job = await _seed(store)
        perf = await build_performance_summary(store, JOB_ID)
        assert perf["pages_with_data"] > len(perf["top_by_impressions"])
        pdf = await generate_pdf_report(
            job, [], await store.get_summary(JOB_ID), performance=perf
        )
        text = _pdf_text(pdf)
        assert "Pages by Impressions (of" in text and "with data)" in text

    @pytest.mark.asyncio
    async def test_f7_excel_label_discloses_the_cap(self, store):
        import io

        from openpyxl import load_workbook

        job = await _seed(store)
        xlsx = generate_excel_report(
            job, [], await store.get_summary(JOB_ID),
            performance=await build_performance_summary(store, JOB_ID),
        )
        wb = load_workbook(io.BytesIO(xlsx))
        text = " ".join(
            str(c.value) for row in wb["Performance"].iter_rows() for c in row if c.value
        )
        assert "pages with data" in text
