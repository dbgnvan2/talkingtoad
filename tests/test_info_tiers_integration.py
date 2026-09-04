"""Info tiers — API contract (every field the frontend reads, asserted here first).

Spec:    docs/pending/2026-09-01_info-tiers.md §9
Tests:   this file. Unit + scoring: tests/test_info_tiers.py

The test that matters most is ``test_summary_score_follows_info_detail`` next
to ``test_summary_stored_counts_unchanged``: the score MUST move with the
level (the owner's decision) while the stored counts MUST NOT — the pair is
what keeps a higher score from passing as a cleaner site (P31).
"""

from __future__ import annotations

import io
from datetime import datetime, timezone
from unittest.mock import AsyncMock, patch

import pytest

from api.models.issue import Issue
from api.models.job import CrawlJob, CrawlSettings
from api.models.page import CrawledPage

BASE = "https://example.com"
PAGE = f"{BASE}/p"

# (code, category, severity, impact) — one row per tier plus a warning.
_ROWS = [
    ("H1_MISSING", "heading", "warning", 4),
    ("IMG_ALT_MISSING", "image", "info", 3),        # high / Key
    ("META_DESC_MISSING", "metadata", "info", 2),   # medium / Notable
    ("TITLE_TOO_SHORT", "metadata", "info", 1),     # low
    ("REDIRECT_TRAILING_SLASH", "redirect", "info", 0),  # low
]


async def _job(store, job_id="j1", info_detail="all", *, url=PAGE, started_at=None):
    job = CrawlJob(job_id=job_id, target_url=BASE, status="complete", pages_crawled=1,
                   settings=CrawlSettings(info_detail=info_detail),
                   started_at=started_at or datetime.now(timezone.utc))
    await store.create_job(job)
    await store.save_pages([CrawledPage(job_id=job_id, url=url, status_code=200, title="t",
                                        crawled_at=datetime.now(timezone.utc))])
    await store.save_issues([
        Issue(job_id=job_id, page_url=url, category=cat, severity=sev, issue_code=code,
              description=code, recommendation="fix", impact=imp)
        for code, cat, sev, imp in _ROWS
    ])
    return job


# ── POST /start, GET /{job_id}, rescan ────────────────────────────────────


class TestSettingTravelsWithTheJob:
    async def test_start_accepts_info_detail(self, api_client, auth_headers, test_store):
        with patch("api.routers.crawl._run_crawl_background", new_callable=AsyncMock):
            r = await api_client.post("/api/crawl/start", headers=auth_headers,
                                      json={"target_url": BASE, "settings": {"info_detail": "notable"}})
        assert r.status_code == 202, r.text
        job = await test_store.get_job(r.json()["job_id"])
        assert job.settings.info_detail == "notable"

    async def test_start_rejects_bad_info_detail(self, api_client, auth_headers):
        r = await api_client.post("/api/crawl/start", headers=auth_headers,
                                  json={"target_url": BASE, "settings": {"info_detail": "some"}})
        assert r.status_code == 422

    async def test_default_is_all(self):
        assert CrawlSettings().info_detail == "all"
        assert CrawlSettings.model_validate({}).info_detail == "all"  # legacy settings blob

    async def test_job_echoes_info_detail(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="key")
        r = await api_client.get("/api/crawl/j1", headers=auth_headers)
        assert r.status_code == 200
        assert r.json()["settings"]["info_detail"] == "key"

    async def test_rescan_inherits_info_detail(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="notable")
        with patch("api.routers.crawl._run_crawl_background", new_callable=AsyncMock):
            r = await api_client.post("/api/crawl/j1/rescan", headers=auth_headers)
        assert r.status_code == 202, r.text
        new = await test_store.get_job(r.json()["job_id"])
        assert new.settings.info_detail == "notable"


# ── GET /summary (via /results) ───────────────────────────────────────────


class TestSummary:
    async def _summary(self, api_client, auth_headers, job_id):
        r = await api_client.get(f"/api/crawl/{job_id}/results", headers=auth_headers)
        assert r.status_code == 200, r.text
        return r.json()["summary"]

    async def test_summary_score_follows_info_detail(self, api_client, auth_headers, test_store):
        """The owner's decision: the level moves the score. Monotone in tightness."""
        scores = {}
        for n, level in enumerate(("all", "notable", "key", "none")):
            await _job(test_store, job_id=f"j{n}", info_detail=level)
            scores[level] = (await self._summary(api_client, auth_headers, f"j{n}"))["health_score"]
        # deductions: all=10, notable=9, key=7, none=4 (the warning stays)
        assert scores == {"all": 90, "notable": 91, "key": 93, "none": 96}, scores

    async def test_summary_stored_counts_unchanged(self, api_client, auth_headers, test_store):
        """by_severity is what was FOUND; the level changes what is COUNTED."""
        await _job(test_store, job_id="j1", info_detail="all")
        await _job(test_store, job_id="j2", info_detail="none")
        s1 = await self._summary(api_client, auth_headers, "j1")
        s2 = await self._summary(api_client, auth_headers, "j2")
        assert s1["by_severity"] == s2["by_severity"] == {"critical": 0, "warning": 1, "info": 4}
        assert s1["total_issues"] == s2["total_issues"] == 5

    async def test_summary_reports_the_tier_breakdown(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="key")
        s = await self._summary(api_client, auth_headers, "j1")
        assert s["info_detail"] == "key"
        assert s["info_by_tier"] == {"high": 1, "medium": 1, "low": 2}
        assert s["info_scored"] == 1
        assert s["info_excluded"] == 3
        assert s["info_scored"] + s["info_excluded"] == s["by_severity"]["info"]

    async def test_agent_health_follows_the_same_level(self, test_store):
        """Both scores move under one rule so they cannot disagree about what counts."""
        for n, level in enumerate(("all", "none")):
            jid = f"a{n}"
            job = CrawlJob(job_id=jid, target_url=BASE, status="complete", pages_crawled=1,
                           settings=CrawlSettings(info_detail=level))
            await test_store.create_job(job)
            await test_store.save_pages([CrawledPage(job_id=jid, url=PAGE, status_code=200,
                                                     crawled_at=datetime.now(timezone.utc))])
            await test_store.save_issues([Issue(
                job_id=jid, page_url=PAGE, category="ai_readiness", severity="info",
                issue_code="FAQ_SCHEMA_MISSING", description="", recommendation="", impact=2)])
        assert (await test_store.get_summary("a0"))["agent_health_score"] == 98
        assert (await test_store.get_summary("a1"))["agent_health_score"] == 100


# ── GET /results, /results/{category} ─────────────────────────────────────


class TestResultsLists:
    async def test_results_has_info_tier_and_filtered_report(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="notable")
        r = await api_client.get("/api/crawl/j1/results", headers=auth_headers)
        body = r.json()
        codes = {i["issue_code"] for i in body["issues"]}
        assert codes == {"H1_MISSING", "IMG_ALT_MISSING", "META_DESC_MISSING"}
        for i in body["issues"]:
            assert "info_tier" in i and "scored" in i
            assert i["scored"] is True
        tiers = {i["issue_code"]: i["info_tier"] for i in body["issues"]}
        assert tiers == {"H1_MISSING": None, "IMG_ALT_MISSING": "high", "META_DESC_MISSING": "medium"}
        assert body["info_filtered"] == {"hidden": 2, "by_tier": {"low": 2}, "info_detail": "notable"}
        assert body["pagination"]["total_issues"] == 3, "the pager must count the served list"

    async def test_hidden_plus_shown_equals_stored_total(self, api_client, auth_headers, test_store):
        """P31: a filter that loses a row fails here."""
        for n, level in enumerate(("all", "notable", "key", "none")):
            await _job(test_store, job_id=f"j{n}", info_detail=level)
            body = (await api_client.get(f"/api/crawl/j{n}/results", headers=auth_headers)).json()
            assert (len(body["issues"]) + body["info_filtered"]["hidden"]
                    + body["filtered"]["hidden"]) == len(_ROWS), level

    async def test_at_all_nothing_changes(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="all")
        body = (await api_client.get("/api/crawl/j1/results", headers=auth_headers)).json()
        assert len(body["issues"]) == 5
        assert body["info_filtered"]["hidden"] == 0
        assert all(i["scored"] for i in body["issues"])

    async def test_reveal_shows_excluded_rows_as_unscored(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="key")
        body = (await api_client.get("/api/crawl/j1/results?info_detail=all",
                                     headers=auth_headers)).json()
        assert len(body["issues"]) == 5
        scored = {i["issue_code"]: i["scored"] for i in body["issues"]}
        assert scored["IMG_ALT_MISSING"] is True and scored["H1_MISSING"] is True
        assert scored["META_DESC_MISSING"] is False and scored["TITLE_TOO_SHORT"] is False
        assert body["info_filtered"]["hidden"] == 0
        # The score is a property of the SCAN: revealing does not move it.
        assert body["summary"]["health_score"] == 93

    async def test_reveal_cannot_tighten(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="all")
        body = (await api_client.get("/api/crawl/j1/results?info_detail=key",
                                     headers=auth_headers)).json()
        assert len(body["issues"]) == 5 and body["info_filtered"]["info_detail"] == "all"
        assert body["summary"]["health_score"] == 90

    async def test_category_results_reveal_only(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="notable")
        r = await api_client.get("/api/crawl/j1/results/metadata", headers=auth_headers)
        body = r.json()
        assert [i["issue_code"] for i in body["issues"]] == ["META_DESC_MISSING"]
        assert body["info_filtered"] == {"hidden": 1, "by_tier": {"low": 1}, "info_detail": "notable"}
        r = await api_client.get("/api/crawl/j1/results/metadata?info_detail=all", headers=auth_headers)
        body = r.json()
        assert {i["issue_code"]: i["scored"] for i in body["issues"]} == {
            "META_DESC_MISSING": True, "TITLE_TOO_SHORT": False}

    async def test_severity_info_view_is_filtered_too(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="key")
        body = (await api_client.get("/api/crawl/j1/results?severity=info",
                                     headers=auth_headers)).json()
        assert [i["issue_code"] for i in body["issues"]] == ["IMG_ALT_MISSING"]
        assert body["info_filtered"]["hidden"] == 3


# ── GET /pages/issues, /pages ─────────────────────────────────────────────


class TestPerPage:
    async def test_page_issues_respect_info_detail(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="notable")
        r = await api_client.get(f"/api/crawl/j1/pages/issues?url={PAGE}", headers=auth_headers)
        body = r.json()
        flat = [i for issues in body["by_category"].values() for i in issues]
        assert {i["issue_code"] for i in flat} == {"H1_MISSING", "IMG_ALT_MISSING", "META_DESC_MISSING"}
        assert all("info_tier" in i and i["scored"] for i in flat)
        assert body["total_issues"] == 3
        assert body["info_filtered"]["hidden"] == 2
        assert "redirect" not in body["by_category"], "an emptied category must not linger"

    async def test_page_issues_reveal(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="notable")
        r = await api_client.get(f"/api/crawl/j1/pages/issues?url={PAGE}&info_detail=all",
                                 headers=auth_headers)
        flat = [i for issues in r.json()["by_category"].values() for i in issues]
        assert len(flat) == 5 and sum(1 for i in flat if not i["scored"]) == 2

    async def test_page_grade_follows_info_detail(self, api_client, auth_headers, test_store):
        """By Page's citability column and the page-priority grade use the level."""
        for n, level in enumerate(("all", "none")):
            jid = f"g{n}"
            job = CrawlJob(job_id=jid, target_url=BASE, status="complete", pages_crawled=1,
                           settings=CrawlSettings(info_detail=level))
            await test_store.create_job(job)
            await test_store.save_pages([CrawledPage(job_id=jid, url=PAGE, status_code=200,
                                                     crawled_at=datetime.now(timezone.utc))])
            await test_store.save_issues([Issue(
                job_id=jid, page_url=PAGE, category="ai_readiness", severity="info",
                issue_code="FAQ_SCHEMA_MISSING", description="", recommendation="", impact=2)])
        g_all = (await api_client.get("/api/crawl/g0/pages", headers=auth_headers)).json()
        g_none = (await api_client.get("/api/crawl/g1/pages", headers=auth_headers)).json()
        assert g_all["pages"][0]["citability_grade"] == 98
        assert g_none["pages"][0]["citability_grade"] == 100
        from api.services.page_priority import build_page_priority
        assert (await build_page_priority(test_store, "g0"))[0]["health_score"] == 98
        assert (await build_page_priority(test_store, "g1"))[0]["health_score"] == 100


    async def test_page_counts_follow_the_level(self, api_client, auth_headers, test_store):
        """Sweep finding: By Page listed a page as "5 issues" whose drawer showed one."""
        await _job(test_store, info_detail="none")
        body = (await api_client.get("/api/crawl/j1/pages", headers=auth_headers)).json()
        counts = body["pages"][0]["issue_counts"]
        assert counts == {"total": 1, "critical": 0, "warning": 1, "info": 0, "info_excluded": 4}
        await _job(test_store, job_id="j2", info_detail="notable")
        body = (await api_client.get("/api/crawl/j2/pages", headers=auth_headers)).json()
        assert body["pages"][0]["issue_counts"] == {
            "total": 3, "critical": 0, "warning": 1, "info": 2, "info_excluded": 2}

    async def test_page_counts_unchanged_at_all(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="all")
        body = (await api_client.get("/api/crawl/j1/pages", headers=auth_headers)).json()
        assert body["pages"][0]["issue_counts"] == {
            "total": 5, "critical": 0, "warning": 1, "info": 4, "info_excluded": 0}

    async def test_citation_eligibility_uses_the_job_level(self):
        """Sweep finding: the AI_HIGH_VALUE_UNCITED "healthy page" gate scored at
        `all` whatever the job's level, so a page every surface called healthy
        was silently ineligible."""
        from datetime import date
        from api.models.page import CrawledPage
        from api.routers.citations import derive_citation_issues
        page = CrawledPage(page_id="p1", job_id="j", url=PAGE, status_code=200, word_count=800,
                           ai_citation_count_30d=0, ai_citation_last_updated="2026-08-20")
        # 10 low-tier rows in one category (cap 20) + a warning: 79 at `all`, 96 at `notable`.
        rows = [(f"L{n}", 1, "metadata") for n in range(17)] + [("W", 4, "heading")]
        rows_by_url = {PAGE.rstrip("/"): rows}
        today = date(2026, 9, 1)
        at_all = derive_citation_issues([page], rows_by_url, today, "j")
        at_notable = derive_citation_issues([page], rows_by_url, today, "j", info_detail="notable")
        assert [i.issue_code for i in at_all] == []
        assert [i.issue_code for i in at_notable] == ["AI_HIGH_VALUE_UNCITED"]


# ── Exports ───────────────────────────────────────────────────────────────


class TestExports:
    async def test_csv_export_reflects_the_level(self, api_client, auth_headers, test_store):
        await _job(test_store, info_detail="key")
        r = await api_client.get("/api/crawl/j1/export/csv", headers=auth_headers)
        assert r.status_code == 200
        assert "IMG_ALT_MISSING" in r.text and "TITLE_TOO_SHORT" not in r.text

    async def test_csv_carries_the_tier_column(self, api_client, auth_headers, test_store):
        """Sweep finding: a CSV at `key` was just shorter, with nothing in the file saying so."""
        import csv
        await _job(test_store, info_detail="all")
        r = await api_client.get("/api/crawl/j1/export/csv", headers=auth_headers)
        rows = list(csv.DictReader(io.StringIO(r.text)))
        tiers = {row["issue_code"]: row["info_tier"] for row in rows}
        assert tiers["IMG_ALT_MISSING"] == "high" and tiers["TITLE_TOO_SHORT"] == "low"
        assert tiers["H1_MISSING"] == ""

    async def test_pdf_summary_table_shows_scored_beside_excluded(self):
        from api.services.report_generator import _info_notices_figure
        assert _info_notices_figure({"by_severity": {"info": 5}, "info_scored": 5, "info_excluded": 0}) == 5
        assert _info_notices_figure({"by_severity": {"info": 5}, "info_scored": 2, "info_excluded": 3}) == "2 (+3 excluded)"
        assert _info_notices_figure({"by_severity": {"info": 5}}) == 5  # legacy summary

    async def test_excel_export_caveat_when_info_excluded(self, api_client, auth_headers, test_store):
        from openpyxl import load_workbook
        await _job(test_store, info_detail="notable")
        r = await api_client.get("/api/crawl/j1/export/excel", headers=auth_headers)
        assert r.status_code == 200
        wb = load_workbook(io.BytesIO(r.content))
        cells = [str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value]
        body = " ".join(c for c in cells if "Scored at info detail" not in c)
        assert "TITLE_TOO_SHORT" not in body, "the workbook lists a finding the scan excluded"
        assert "META_DESC_MISSING" in body
        caveat = [c for c in cells if "Scored at info detail" in c]
        assert caveat and "2 info notices" in caveat[0] and "health score" in caveat[0]

    async def test_pdf_export_caveat_when_info_excluded(self, api_client, auth_headers, test_store):
        from pypdf import PdfReader
        await _job(test_store, info_detail="notable")
        r = await api_client.get("/api/crawl/j1/export/pdf", headers=auth_headers)
        assert r.status_code == 200
        text = " ".join((pg.extract_text() or "") for pg in PdfReader(io.BytesIO(r.content)).pages)
        assert "Scored at info detail" in text.replace("\n", " ")

    async def test_no_caveat_at_all(self, api_client, auth_headers, test_store):
        from openpyxl import load_workbook
        await _job(test_store, info_detail="all")
        r = await api_client.get("/api/crawl/j1/export/excel", headers=auth_headers)
        wb = load_workbook(io.BytesIO(r.content))
        cells = " ".join(str(c.value) for ws in wb.worksheets for row in ws.iter_rows() for c in row if c.value)
        assert "Scored at info detail" not in cells


# ── P5.2: a count beside a score must be the population the score charged ──
# The tile/list disagreement below is what P5.2 exists to close. Measured before
# the fix, at info_detail="key": the `metadata` tile read 2 and opened an EMPTY
# list, and the PDF's per-page row reported `info_excluded: 0` on a job that
# excluded three rows — a disclosure field positively asserting the opposite of
# the truth, which is worse than not having one.


class TestCountsMatchTheScoredPopulation:
    _CATS = ("metadata", "redirect", "heading", "image")

    async def _summary(self, api_client, auth_headers, job_id="j1"):
        r = await api_client.get(f"/api/crawl/{job_id}/results", headers=auth_headers)
        assert r.status_code == 200, r.text
        return r.json()["summary"]

    @pytest.mark.parametrize("level", ["all", "notable", "key", "none"])
    async def test_category_tile_count_equals_the_list_it_opens(
        self, api_client, auth_headers, test_store, level
    ):
        """4.1 — read one side, assert against the other side's LIVE output.

        A tile is a button; its number is a promise about what opening it shows.
        This asserts the summary's per-category count against the actual length of
        `/results/{category}` — not against a second computation of the same thing,
        which is how the disagreement shipped green (LEARNINGS checklist 13).
        """
        await _job(test_store, job_id="j1", info_detail=level)
        summary = await self._summary(api_client, auth_headers)
        scored = summary.get("by_category_scored") or {}
        for cat in self._CATS:
            r = await api_client.get(f"/api/crawl/j1/results/{cat}", headers=auth_headers)
            assert r.status_code == 200, r.text
            listed = len(r.json()["issues"])
            assert scored.get(cat, 0) == listed, (
                f"at info_detail={level!r} the {cat} tile says {scored.get(cat, 0)} "
                f"and the list it opens has {listed} rows"
            )

    @pytest.mark.parametrize("level", ["all", "notable", "key", "none"])
    async def test_scored_plus_excluded_equals_stored_per_category(
        self, api_client, auth_headers, test_store, level
    ):
        """4.2 — the same invariant info_scored + info_excluded already carries."""
        await _job(test_store, job_id="j1", info_detail=level)
        s = await self._summary(api_client, auth_headers)
        stored, scored, excluded = (s["by_category"], s["by_category_scored"],
                                    s["by_category_excluded"])
        # Cannot fail while all three are seeded from one comprehension over
        # PHASE_1_CATEGORIES — kept as the guard for the day they are not, and
        # labelled so nobody counts it as coverage. The arithmetic below is the
        # assertion that does work.
        assert set(scored) == set(stored) == set(excluded), "the three maps disagree on keys"
        for cat in stored:
            assert scored[cat] + excluded[cat] == stored[cat], (
                f"{cat}: {scored[cat]} + {excluded[cat]} != {stored[cat]}"
            )

    async def test_by_category_stays_the_stored_count(
        self, api_client, auth_headers, test_store
    ):
        """4.3 — the other direction of 4.2, and the obvious WRONG fix.

        Redefining `by_category` as the scored count makes 4.1 pass with a smaller
        diff, and leaves `by_severity` (stored) and `by_category` (scored) meaning
        different things in one response. `test_summary_stored_counts_unchanged`
        pins the severity half; this pins the category half.
        """
        await _job(test_store, job_id="j1", info_detail="all")
        await _job(test_store, job_id="j2", info_detail="none")
        s1 = await self._summary(api_client, auth_headers, "j1")
        s2 = await self._summary(api_client, auth_headers, "j2")
        assert s1["by_category"] == s2["by_category"], (
            "by_category moved with the level — it is the FOUND count, like by_severity"
        )
        assert s1["by_category"]["metadata"] == 2
        # ...and the scored map is the one that moves.
        assert s1["by_category_scored"]["metadata"] == 2
        assert s2["by_category_scored"]["metadata"] == 0

    async def test_tile_hint_exists_whenever_a_category_lost_rows(
        self, api_client, auth_headers, test_store
    ):
        """4.3b — a scored count with no disclosure is P31 wearing a fix.

        At `none` every tile reads 0, and 0 is what a clean site looks like. The
        per-category excluded map is what the tile's "+N not scored" line renders
        from, so it must be non-zero exactly where rows were dropped.
        """
        await _job(test_store, job_id="j1", info_detail="none")
        s = await self._summary(api_client, auth_headers)
        assert s["by_category_scored"]["metadata"] == 0
        assert s["by_category_excluded"]["metadata"] == 2, (
            "the tile would render a bare 0 with nothing saying rows were excluded"
        )


class TestPageCountCallersPassTheLevel:
    async def test_pdf_page_rows_use_the_jobs_level(
        self, api_client, auth_headers, test_store
    ):
        """4.4 — the PDF's "Pages with Most Issues" table."""
        from pypdf import PdfReader
        await _job(test_store, info_detail="key")
        r = await api_client.get("/api/crawl/j1/export/pdf", headers=auth_headers)
        assert r.status_code == 200
        text = " ".join((pg.extract_text() or "") for pg in
                        PdfReader(io.BytesIO(r.content)).pages).replace("\n", " ")
        assert "4 Info" not in text, (
            "the per-page row printed the stored info count (4); the scan scored 1"
        )
        assert "1 Info" in text

    async def test_pdf_page_row_prints_what_the_level_took_off_that_page(
        self, api_client, auth_headers, test_store
    ):
        """4.4b — assert the RENDERED artifact, not the number behind it.

        4.4 checks the count is scoped and 4.5 checks the store's disclosure
        field; deleting the PDF's per-page "(+N excluded)" line left both green.
        The row must SAY what it dropped, or a page whose findings were all
        excluded prints as a clean page.

        A second page is seeded so the per-page figures (3 and 2) differ from the
        site total (5) — otherwise the Dashboard's "Info Notices: 1 (+3 excluded)"
        line satisfies the assertion on its own and the test proves nothing.
        """
        from pypdf import PdfReader

        await _job(test_store, info_detail="key")
        other = f"{BASE}/other"
        await test_store.save_pages([CrawledPage(
            job_id="j1", url=other, status_code=200, title="o",
            crawled_at=datetime.now(timezone.utc))])
        await test_store.save_issues([
            Issue(job_id="j1", page_url=other, category=cat, severity=sev,
                  issue_code=code, description=code, recommendation="fix", impact=imp)
            for code, cat, sev, imp in _ROWS[3:]      # 2 info rows, both excluded at `key`
        ])

        r = await api_client.get("/api/crawl/j1/export/pdf", headers=auth_headers)
        assert r.status_code == 200
        text = " ".join((pg.extract_text() or "") for pg in
                        PdfReader(io.BytesIO(r.content)).pages).replace("\n", " ")
        assert "(+3 excluded)" in text, "the page row did not say what the level dropped"
        assert "(+2 excluded)" in text, (
            "only the first page's row carried the disclosure — and this figure "
            "cannot come from the site-total line, which says 5"
        )

    async def test_pdf_page_rows_never_claim_zero_excluded_when_rows_were(
        self, test_store
    ):
        """4.5 — the disclosure figure, against an oracle from the FIXTURE.

        Honest about what this is: the count and the disclosure are complementary
        branches of one CASE in one SELECT, so they cannot diverge by
        construction, and an earlier docstring here claimed otherwise. What it
        does pin is the figure itself, computed from `_ROWS` and
        `info_row_excluded` — the catalogue's own predicate, not the query's — so
        it fails if the SQL and the Python part ways. That is not hypothetical:
        `_kept_info_sql` was missing `info_row_excluded`'s "impact >= 4 is never
        excluded" clause until 2026-09-04, and thousands of live rows hit it.
        """
        from api.crawler.checkers.registry import info_row_excluded

        await _job(test_store, job_id="j1", info_detail="key")
        expected_excluded = sum(1 for _c, _cat, sev, imp in _ROWS
                                if sev == "info" and info_row_excluded(imp, "key"))
        expected_kept = len(_ROWS) - expected_excluded

        pages, _ = await test_store.get_pages_with_issue_counts("j1", info_detail="key")
        counts = pages[0]["issue_counts"]
        assert counts["total"] == expected_kept, counts
        assert counts["info_excluded"] == expected_excluded, (
            f"the row claims {counts['info_excluded']} rows were excluded; "
            f"info_row_excluded says {expected_excluded}"
        )
        assert counts["info_excluded"] > 0, "the fixture stopped exercising exclusion"

    async def test_advisor_membership_covers_the_whole_crawl_not_the_top_50(
        self, api_client, auth_headers, test_store
    ):
        """4.6c — the P9 underneath, surfaced by passing the level.

        `get_pages_with_issue_counts` defaults to `limit=50` and orders by issue
        count, so a bare call answers "is this one of the 50 worst pages", not
        "is this page in the job". Two consequences: a 500-page crawl rejected
        90% of its own URLs, and passing info_detail changed `total`, hence the
        ordering, hence which URLs validated — so the fix for one defect silently
        moved the boundary of another.

        A clean page (no issues at all) sorts LAST, which is exactly the page a
        limited window drops.
        """
        await _job(test_store, info_detail="key")
        clean = f"{BASE}/page-with-nothing-wrong"
        await test_store.save_pages(
            [CrawledPage(job_id="j1", url=f"{BASE}/filler-{n}", status_code=200,
                         title="f", crawled_at=datetime.now(timezone.utc))
             for n in range(60)]
            + [CrawledPage(job_id="j1", url=clean, status_code=200, title="c",
                           crawled_at=datetime.now(timezone.utc))]
        )
        # Give the filler pages an issue each so they outrank the clean page.
        await test_store.save_issues([
            Issue(job_id="j1", page_url=f"{BASE}/filler-{n}", category="heading",
                  severity="warning", issue_code="H1_MISSING", description="d",
                  recommendation="fix", impact=4)
            for n in range(60)
        ])

        with patch("api.routers.advisor.evaluate_page",
                   AsyncMock(return_value=("# report", False))):
            r = await api_client.post(
                "/api/ai/geo-report",
                json={"job_id": "j1", "page_urls": [clean]},
                headers=auth_headers,
            )
        assert r.status_code != 400, (
            "a page in the job was rejected as not in the job — membership was "
            f"answered over a truncated window: {r.text}"
        )
        assert r.status_code == 200, r.text

    def test_every_caller_of_get_pages_with_issue_counts_passes_info_detail(self):
        """4.6 — structural, and it says so.

        This proves the ARGUMENT is passed at every call site, not that the value
        is right — the behavioural tests above cover the callers that exist today.
        The next caller is the one that will forget: three of the four omitted it
        before P5.2, which is how the PDF and the advisor came to disagree with
        the score.

        What it cannot see, stated so nobody reads more into a green: a call made
        through an alias (`fn = store.get_pages_with_issue_counts`), or from
        outside `api/`. A `**kwargs` splat fails closed, which is the right way
        round.
        """
        import re
        from pathlib import Path

        def _strip_comments(text: str) -> str:
            """A `# info_detail: deliberately omitted` comment inside the call
            satisfied a bare substring check — verified, it did."""
            return re.sub(r"#[^\n]*", "", text)

        api = Path(__file__).resolve().parent.parent / "api"
        offenders = []
        seen_defs = 0
        for path in api.rglob("*.py"):
            src = path.read_text()
            for m in re.finditer(r"(def\s+)?get_pages_with_issue_counts\s*\(", src):
                if m.group(1):          # the definition itself
                    seen_defs += 1
                    continue
                # Balance parens from the call's opening bracket.
                i, depth = m.end() - 1, 0
                while i < len(src):
                    if src[i] == "(":
                        depth += 1
                    elif src[i] == ")":
                        depth -= 1
                        if depth == 0:
                            break
                    i += 1
                if "info_detail" not in _strip_comments(src[m.end():i]):
                    line = src[:m.start()].count("\n") + 1
                    offenders.append(f"{path.relative_to(api.parent)}:{line}")
        assert seen_defs == 1, (
            f"expected exactly one definition to skip, found {seen_defs} — the "
            "scan's shape has changed and its results are not trustworthy"
        )
        assert not offenders, (
            "get_pages_with_issue_counts called without info_detail (the default is "
            f"'all', which silently reports stored counts): {offenders}"
        )

    async def test_advisor_page_list_ranks_by_the_scored_count(
        self, api_client, auth_headers, test_store
    ):
        """4.6b — the picker RANKS by this number, so a stored count promotes a
        page whose findings the scan excluded: offered first, opens empty.

        Two pages, because the first version of this test had one and there was
        no order to get wrong. The second carries more STORED rows and fewer
        SCORED ones, so the stored and scored orderings are opposites and the
        assertion can only be satisfied one way.

        Where the order comes from is deliberately not asserted: the store
        already returns `ORDER BY total DESC` and the router's `out.sort` is
        belt-and-braces, so deleting either one alone leaves this green. What is
        pinned is the property the operator sees — the page worth fixing is
        offered first — not which layer produces it.
        """
        await _job(test_store, job_id="j1", info_detail="key")
        noisy = f"{BASE}/all-low-impact"
        await test_store.save_pages([CrawledPage(
            job_id="j1", url=noisy, status_code=200, title="n",
            crawled_at=datetime.now(timezone.utc))])
        await test_store.save_issues([
            Issue(job_id="j1", page_url=noisy, category="metadata", severity="info",
                  issue_code=f"TITLE_TOO_SHORT_{n}", description="d",
                  recommendation="fix", impact=1)
            for n in range(9)          # 9 stored rows, 0 scored at `key`
        ])

        r = await api_client.get("/api/ai/geo-report/pages?job_id=j1", headers=auth_headers)
        assert r.status_code == 200, r.text
        rows = r.json()["pages"]
        counts = {p["url"]: p["issue_count"] for p in rows}
        assert counts[PAGE] == 2, f"the picker counted stored rows: {counts}"
        assert counts[noisy] == 0, f"the picker counted stored rows: {counts}"
        assert [p["url"] for p in rows][0] == PAGE, (
            "ranked on the stored count — the page with 9 excluded rows and "
            f"nothing to fix was offered first: {[p['url'] for p in rows]}"
        )


class TestPdfAndExcelTotals:
    async def test_pdf_total_issues_shows_scored_when_any_excluded(
        self, api_client, auth_headers, test_store
    ):
        """4.7 — the frontend already does this (SummaryPanel "found · N scored").

        The PDF printed the bare stored number two rows under Health Score. P16:
        a capability added at one front end only.
        """
        from pypdf import PdfReader
        await _job(test_store, info_detail="key")
        r = await api_client.get("/api/crawl/j1/export/pdf", headers=auth_headers)
        text = " ".join((pg.extract_text() or "") for pg in
                        PdfReader(io.BytesIO(r.content)).pages).replace("\n", " ")
        assert "5 (2 scored)" in text, (
            "Total Issues Found printed a stored count with nothing saying so"
        )

    async def test_pdf_total_is_bare_when_nothing_excluded(
        self, api_client, auth_headers, test_store
    ):
        """4.7b — the other direction: no parenthetical on a full-detail scan."""
        from pypdf import PdfReader
        await _job(test_store, info_detail="all")
        r = await api_client.get("/api/crawl/j1/export/pdf", headers=auth_headers)
        text = " ".join((pg.extract_text() or "") for pg in
                        PdfReader(io.BytesIO(r.content)).pages).replace("\n", " ")
        assert "scored)" not in text

    async def test_pdf_category_table_uses_the_scored_count(
        self, api_client, auth_headers, test_store
    ):
        """4.9 — the PDF's OWN "Issues by Category" table, missed by the first
        pass of P5.2 and found by the cold sweep.

        With only the Excel and tile surfaces fixed, one audit's two exports
        disagreed: the workbook omitted Metadata entirely while the PDF printed
        "Metadata: 2" three pages before a findings list containing no metadata
        rows, and two rows under a "Total Issues Found: 5 (2 scored)" the same
        commit had just made honest.
        """
        from pypdf import PdfReader
        await _job(test_store, info_detail="key")
        r = await api_client.get("/api/crawl/j1/export/pdf", headers=auth_headers)
        assert r.status_code == 200
        text = " ".join((pg.extract_text() or "") for pg in
                        PdfReader(io.BytesIO(r.content)).pages).replace("\n", " ")
        assert "Metadata: 2" not in text, (
            "the category table printed the stored count beside a scoped score"
        )
        assert "Metadata: 0 (2 not scored)" in text, (
            "a bare 0 is what a clean category looks like — the caveat travels "
            "with the count on every surface, not just the tiles"
        )
        assert "Headings: 1" in text, "the table lost a category that did score"

    async def test_excel_category_sheet_uses_the_scored_count(
        self, api_client, auth_headers, test_store
    ):
        """4.8 — the workbook already carries the caveat sentence; the number
        beside it was the stored one."""
        from openpyxl import load_workbook

        def _cats(job_id):
            wb = load_workbook(io.BytesIO(_xl[job_id]))
            rows = [[c.value for c in row] for ws in wb.worksheets for row in ws.iter_rows()]
            return {r[0]: r[1] for r in rows
                    if r and isinstance(r[0], str) and len(r) > 1 and isinstance(r[1], int)}

        _xl = {}
        for jid, level in (("j1", "key"), ("j2", "all")):
            await _job(test_store, job_id=jid, info_detail=level)
            r = await api_client.get(f"/api/crawl/{jid}/export/excel", headers=auth_headers)
            assert r.status_code == 200
            _xl[jid] = r.content

        # At `key` both metadata rows are excluded, so the count is 0 — and the
        # row STAYS, carrying what was left out. Dropping it (the sheet omits
        # zero rows) would take the disclosure with it, and a category absent
        # from the table is indistinguishable from a category with no findings.
        assert _cats("j1").get("Metadata") == 0, (
            f"the category sheet still counts the excluded metadata rows: {_cats('j1')}"
        )
        assert _cats("j1").get("Heading") == 1, "the category table lost its real rows"
        # ...and the other direction: at `all` nothing changes and no caveat appears.
        assert _cats("j2").get("Metadata") == 2, "at info_detail=all nothing should change"

        def _notes(job_id):
            wb = load_workbook(io.BytesIO(_xl[job_id]))
            return [str(c.value) for ws in wb.worksheets for row in ws.iter_rows()
                    for c in row if c.value and "not scored at info detail" in str(c.value)]

        assert _notes("j1"), "the zero came with nothing saying rows were excluded"
        # Not [0] — the notes are in category order, so metadata is not first.
        assert any("2 not scored" in n for n in _notes("j1")), _notes("j1")
        assert not _notes("j2"), "a full-detail scan must carry no per-category caveat"


# ── GET /comparison ───────────────────────────────────────────────────────


class TestComparison:
    async def test_compare_flags_info_detail_mismatch(self, api_client, auth_headers, test_store):
        old = datetime(2026, 1, 1, tzinfo=timezone.utc)
        await _job(test_store, job_id="old", info_detail="all", started_at=old)
        await _job(test_store, job_id="new", info_detail="notable")
        body = (await api_client.get("/api/crawl/new/comparison", headers=auth_headers)).json()
        assert body["comparison_available"] is True
        assert body["comparable"] is False
        assert body["reason"] == "info_detail differs (notable vs all)"
        assert body["current"]["info_detail"] == "notable"
        assert body["previous"]["info_detail"] == "all"
        assert body["delta"]["health_score"] == 1  # still returned, struck through by the UI

    async def test_compare_carries_both_score_bases_and_flags_a_partial_scan(self, api_client, auth_headers, test_store):
        """Phase 4 U4.2 — the compare card strikes the delta through for a partial analysis too."""
        old = datetime(2026, 1, 1, tzinfo=timezone.utc)
        await _job(test_store, job_id="old", info_detail="all", started_at=old)
        job = CrawlJob(job_id="new", target_url=BASE, status="complete", pages_crawled=1,
                       settings=CrawlSettings(enabled_analyses=["link_integrity"]),
                       started_at=datetime.now(timezone.utc))
        await test_store.create_job(job)
        await test_store.save_pages([CrawledPage(job_id="new", url=PAGE, status_code=200,
                                                 crawled_at=datetime.now(timezone.utc))])
        body = (await api_client.get("/api/crawl/new/comparison", headers=auth_headers)).json()
        assert "health_score_basis" in body["current"] and "health_score_basis" in body["previous"]
        assert body["comparable"] is False and "partial analysis" in body["reason"]

    async def test_compare_same_level_is_comparable(self, api_client, auth_headers, test_store):
        old = datetime(2026, 1, 1, tzinfo=timezone.utc)
        await _job(test_store, job_id="old", info_detail="key", started_at=old)
        await _job(test_store, job_id="new", info_detail="key")
        body = (await api_client.get("/api/crawl/new/comparison", headers=auth_headers)).json()
        assert body["comparable"] is True and body["reason"] is None
